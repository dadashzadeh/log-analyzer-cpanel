import re
import json
import gzip
import zipfile
import tarfile
import ipaddress
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
import requests
from user_agents import parse
import socket
import os
from pathlib import Path
import urllib.parse
import base64
import hashlib
from typing import Dict, List, Tuple, Set, Any

class GoogleBotVerifier:
    """کلاس تایید بات‌های گوگل با فایل‌های رسمی"""
    
    def __init__(self):
        self.google_ip_ranges = {
            'googlebot': [],
            'special_crawlers': [],
            'user_triggered': [],
            'user_triggered_google': []
        }
        self.load_google_ip_files()
        
    def load_google_ip_files(self):
        """بارگذاری فایل‌های JSON گوگل"""
        json_files = {
            'googlebot': 'googlebot.json',
            'special_crawlers': 'special-crawlers.json',
            'user_triggered': 'user-triggered-fetchers.json',
            'user_triggered_google': 'user-triggered-fetchers-google.json'
        }
        
        for key, filename in json_files.items():
            if os.path.exists(filename):
                try:
                    with open(filename, 'r') as f:
                        data = json.load(f)
                        if 'prefixes' in data:
                            for prefix in data['prefixes']:
                                if 'ipv4Prefix' in prefix:
                                    self.google_ip_ranges[key].append(
                                        ipaddress.IPv4Network(prefix['ipv4Prefix'])
                                    )
                                elif 'ipv6Prefix' in prefix:
                                    self.google_ip_ranges[key].append(
                                        ipaddress.IPv6Network(prefix['ipv6Prefix'])
                                    )
                    print(f"✅ {filename}: {len(self.google_ip_ranges[key])} رنج IP")
                except Exception as e:
                    print(f"⚠️ خطا در {filename}: {e}")
    
    def verify_google_bot(self, ip_str, user_agent):
        """تایید هویت بات گوگل"""
        result = {
            'is_google': False,
            'bot_type': None,
            'verification_method': None,
            'details': {}
        }
        
        ua_lower = user_agent.lower()
        google_ua_patterns = {
            'googlebot': ['googlebot', 'mediapartners-google', 'google-inspectiontool'],
            'adsbot': ['adsbot-google', 'adsbot-google-mobile'],
            'other': ['google-site-verification', 'chrome-lighthouse', 'googleother']
        }
        
        detected_bot = None
        for bot_type, patterns in google_ua_patterns.items():
            if any(pattern in ua_lower for pattern in patterns):
                detected_bot = bot_type
                break
        
        if not detected_bot:
            return result
        
        result['bot_type'] = detected_bot
        
        try:
            ip_obj = ipaddress.ip_address(ip_str)
            
            for range_type, ip_networks in self.google_ip_ranges.items():
                for network in ip_networks:
                    if ip_obj in network:
                        result['is_google'] = True
                        result['verification_method'] = f'Google {range_type} IP range'
                        result['details']['ip_range'] = str(network)
                        break
                if result['is_google']:
                    break
            
            if not result['is_google']:
                try:
                    hostname = socket.gethostbyaddr(ip_str)[0]
                    result['details']['hostname'] = hostname
                    
                    google_dns_patterns = [
                        r'.*\.googlebot\.com$',
                        r'.*\.google\.com$',
                        r'.*\.googleusercontent\.com$'
                    ]
                    
                    for pattern in google_dns_patterns:
                        if re.match(pattern, hostname):
                            forward_ips = socket.gethostbyname_ex(hostname)[2]
                            if ip_str in forward_ips:
                                result['is_google'] = True
                                result['verification_method'] = 'DNS verification'
                                break
                except:
                    pass
        except:
            pass
        
        return result

class AdvancedSecurityAnalyzer:
    """تحلیلگر امنیتی پیشرفته با قابلیت‌های جامع"""
    
    def __init__(self, log_file_path: str, site_type: str = 'general'):
        self.log_file_path = log_file_path
        self.site_type = site_type.lower()
        self.logs = []
        self.suspicious_ips = set()
        self.critical_ips = set()  # IP های با خطر بالا
        self.fake_bots = defaultdict(dict)
        self.analysis_results = {}

        # بات‌های معتبر
        self.legitimate_bots = {
            'Google': {
                'patterns': ['googlebot', 'adsbot-google', 'mediapartners-google', 'google-inspectiontool'],
                'dns_suffix': ['.googlebot.com', '.google.com', '.googleusercontent.com'],
                'ip_ranges': []
            },
            'Bing': {
                'patterns': ['bingbot', 'msnbot', 'bingpreview'],
                'dns_suffix': ['.search.msn.com'],
                'ip_ranges': []
            },
            'OpenAI': {
                'patterns': ['gptbot', 'oai-searchbot', 'chatgpt-user', 'openai', 'chatgpt'],
                'dns_suffix': ['.openai.com'],
                'ip_ranges': []
            },
            'PerplexityBot': {
                'patterns': ['perplexitybot', 'perplexity'],
                'dns_suffix': ['.perplexity.ai'],
                'ip_ranges': []
            },
            'PerplexityUser': {
                'patterns': ['perplexity-user'],
                'dns_suffix': ['.perplexity.ai'],
                'ip_ranges': []
            },
            'GoogleCloud': {
                'patterns': [],
                'dns_suffix': ['.google.com'],
                'ip_ranges': []
            },
            'Meta': {
                'patterns': ['facebookbot', 'facebookexternalhit', 'meta-externalagent', 'meta-externalfetcher', 'meta-externalads', 'facebookcatalog'],
                'dns_suffix': ['.facebook.com'],
                'ip_ranges': []
            },
            'LinkedIn': {
                'patterns': ['linkedinbot'],
                'dns_suffix': ['.linkedin.com'],
                'ip_ranges': []
            },
            'ByteDance': {
                'patterns': ['bytespider'],
                'dns_suffix': ['.bytedance.com'],
                'ip_ranges': []
            },
            'DuckDuckGo': {
                'patterns': ['duckassistbot'],
                'dns_suffix': ['.duckduckgo.com'],
                'ip_ranges': []
            },
            'Cohere': {
                'patterns': ['cohere-ai'],
                'dns_suffix': ['.cohere.ai'],
                'ip_ranges': []
            },
            'AllenInstitute': {
                'patterns': ['ai2bot'],
                'dns_suffix': ['.allenai.org'],
                'ip_ranges': []
            },
            'CommonCrawl': {
                'patterns': ['ccbot'],
                'dns_suffix': ['.commoncrawl.org'],
                'ip_ranges': []
            },
            'Diffbot': {
                'patterns': ['diffbot'],
                'dns_suffix': ['.diffbot.com'],
                'ip_ranges': []
            },
            'Omgili': {
                'patterns': ['omgili'],
                'dns_suffix': ['.omgili.com'],
                'ip_ranges': []
            },
            'Timpi': {
                'patterns': ['timpi', 'timpipbot'],
                'dns_suffix': ['.timpi.io'],
                'ip_ranges': []
            },
            'YouCom': {
                'patterns': ['youbot'],
                'dns_suffix': ['.you.com'],
                'ip_ranges': []
            },
            'Mistral': {
                'patterns': ['mistralai-user'],
                'dns_suffix': ['.mistral.ai'],
                'ip_ranges': []
            },
            'Amazon': {
                'patterns': ['amazonbot'],
                'dns_suffix': ['.amazon.com'],
                'ip_ranges': []
            },
            'Apple': {
                'patterns': ['applebot', 'applebot-extended'],
                'dns_suffix': ['.apple.com', '.applebot.apple.com'],
                'ip_ranges': []
            },
            'Yandex': {
                'patterns': ['yandexbot'],
                'dns_suffix': ['.yandex.ru', '.yandex.net'],
                'ip_ranges': []
            },
            'Baidu': {
                'patterns': ['baiduspider'],
                'dns_suffix': ['.baidu.com'],
                'ip_ranges': []
            },
            'SemRush': {
                'patterns': ['semrushbot'],
                'dns_suffix': ['.semrush.com'],
                'ip_ranges': []
            },
            'Ahrefs': {
                'patterns': ['ahrefsbot'],
                'dns_suffix': ['.ahrefs.com'],
                'ip_ranges': []
            },
            'Google-Extended': {
                'patterns': ['google-extended'],
                'dns_suffix': ['.google.com'],
                'ip_ranges': []
            },
            'Google-CloudVertexBot': {
                'patterns': ['google-cloudvertexbot'],
                'dns_suffix': ['.google.com'],
                'ip_ranges': []
            },
            'GoogleOther': {
                'patterns': ['googleother'],
                'dns_suffix': ['.google.com'],
                'ip_ranges': []
            }
        }
        

        # Google Bot Verifier
        self.google_verifier = GoogleBotVerifier()
        for range_type, ip_networks in self.google_verifier.google_ip_ranges.items():
            for network in ip_networks:
                self.legitimate_bots['Google']['ip_ranges'].append(str(network))

        
        # بارگذاری IPهای Bing از فایل bingbot.json
        bingbot_file = 'bingbot.json'
        if os.path.exists(bingbot_file):
            try:
                with open(bingbot_file, 'r') as f:
                    data = json.load(f)
                    if 'prefixes' in data:
                        for prefix in data['prefixes']:
                            if 'ipv4Prefix' in prefix:
                                self.legitimate_bots['Bing']['ip_ranges'].append(prefix['ipv4Prefix'])
                            elif 'ipv6Prefix' in prefix:
                                self.legitimate_bots['Bing']['ip_ranges'].append(prefix['ipv6Prefix'])
                print(f"✅ bingbot.json: {len(self.legitimate_bots['Bing']['ip_ranges'])} رنج IP")
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری bingbot.json: {e}")

        # بارگذاری IPهای OpenAI از فایل gptbot.json
        openai_file = 'gptbot.json'
        if os.path.exists(openai_file):
            try:
                with open(openai_file, 'r') as f:
                    data = json.load(f)
                    if 'prefixes' in data:
                        for prefix in data['prefixes']:
                            if 'ipv4Prefix' in prefix:
                                self.legitimate_bots['OpenAI']['ip_ranges'].append(prefix['ipv4Prefix'])
                            elif 'ipv6Prefix' in prefix:
                                self.legitimate_bots['OpenAI']['ip_ranges'].append(prefix['ipv6Prefix'])
                print(f"✅ gptbot.json: {len(self.legitimate_bots['OpenAI']['ip_ranges'])} رنج IP")
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری gptbot.json: {e}")

        # بارگذاری IPهای Perplexity Bot از فایل perplexitybot.json
        perplexity_bot_file = 'perplexitybot.json'
        if os.path.exists(perplexity_bot_file):
            try:
                with open(perplexity_bot_file, 'r') as f:
                    data = json.load(f)
                    if 'prefixes' in data:
                        for prefix in data['prefixes']:
                            if 'ipv4Prefix' in prefix:
                                self.legitimate_bots['PerplexityBot']['ip_ranges'].append(prefix['ipv4Prefix'])
                            elif 'ipv6Prefix' in prefix:
                                self.legitimate_bots['PerplexityBot']['ip_ranges'].append(prefix['ipv6Prefix'])
                print(f"✅ perplexitybot.json: {len(self.legitimate_bots['PerplexityBot']['ip_ranges'])} رنج IP")
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری perplexitybot.json: {e}")

        # بارگذاری IPهای Perplexity User از فایل perplexity-user.json
        perplexity_user_file = 'perplexity-user.json'
        if os.path.exists(perplexity_user_file):
            try:
                with open(perplexity_user_file, 'r') as f:
                    data = json.load(f)
                    if 'prefixes' in data:
                        for prefix in data['prefixes']:
                            if 'ipv4Prefix' in prefix:
                                self.legitimate_bots['PerplexityUser']['ip_ranges'].append(prefix['ipv4Prefix'])
                            elif 'ipv6Prefix' in prefix:
                                self.legitimate_bots['PerplexityUser']['ip_ranges'].append(prefix['ipv6Prefix'])
                print(f"✅ perplexity-user.json: {len(self.legitimate_bots['PerplexityUser']['ip_ranges'])} رنج IP")
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری perplexity-user.json: {e}")

        # بارگذاری IPهای Google Cloud از فایل cloud.json
        cloud_file = 'cloud.json'
        if os.path.exists(cloud_file):
            try:
                with open(cloud_file, 'r') as f:
                    data = json.load(f)
                    if 'prefixes' in data:
                        for prefix in data['prefixes']:
                            if 'ipv4Prefix' in prefix:
                                self.legitimate_bots['GoogleCloud']['ip_ranges'].append(prefix['ipv4Prefix'])
                            elif 'ipv6Prefix' in prefix:
                                self.legitimate_bots['GoogleCloud']['ip_ranges'].append(prefix['ipv6Prefix'])
                print(f"✅ cloud.json: {len(self.legitimate_bots['GoogleCloud']['ip_ranges'])} رنج IP")
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری cloud.json: {e}")

        # تنظیمات امنیتی
        self.security_thresholds = {
            'requests_per_minute': 30,
            'requests_per_hour': 500,
            'failed_login_attempts': 5,
            'error_404_threshold': 20,
            'large_response_size': 1048576,  # 1MB
            'suspicious_score_threshold': 50
        }

        # الگوهای حملات پیشرفته
        self.advanced_attack_patterns = {
            'sql_injection': {
                'patterns': [
                    r'(union|select|insert|update|delete|drop)[\s\+]+',
                    r'(or|and)[\s\+]*[\'\"\d][\s\+]*=',
                    r'exec(\s|\+|%20)*\((sp_|xp_)',
                    r'(;|%3B)[\s]*(drop|delete|truncate|update)',
                    r'(benchmark|sleep|waitfor[\s]+delay)[\s]*\(',
                    r'(load_file|into[\s]+outfile|into[\s]+dumpfile)',
                    r'concat[\s]*\(.*,.*\)',
                    r'group[\s]+by[\s]+\d+[\s]+having',
                    r'information_schema\.',
                    r'sysobjects|syscolumns|systables'
                ],
                'severity': 'CRITICAL'
            },
            'xss': {
                'patterns': [
                    r'<script[\s>]',
                    r'javascript:',
                    r'on\w+[\s]*=[\s]*["\']',
                    r'<iframe[\s>]',
                    r'<embed[\s>]',
                    r'<object[\s>]',
                    r'document\.(cookie|write|domain|referrer)',
                    r'window\.(location|open)',
                    r'eval[\s]*\(',
                    r'expression[\s]*\(',
                    r'<svg[\s]+on\w+',
                    r'data:text/html',
                    r'vbscript:',
                    r'<img[\s]+src[\s]*=.*on\w+',
                    r'&#x?[0-9a-f]+;'
                ],
                'severity': 'HIGH'
            },
            'lfi_rfi': {
                'patterns': [
                    r'\.\./\.\./\.\./\.\./\.\./etc/passwd',
                    r'\.\.[\\/]',
                    r'%2e%2e[\\/]',
                    r'%252e%252e',
                    r'(include|require|include_once|require_once)[\s]*\(',
                    r'php://input',
                    r'php://filter',
                    r'data://text/plain',
                    r'expect://\w+',
                    r'zip://.*#',
                    r'phar://',
                    r'file:///\w+',
                    r'/proc/self/environ',
                    r'/var/log/\w+',
                    r'c:\\\\windows\\\\',
                    r'c:\\\\winnt\\\\'
                ],
                'severity': 'CRITICAL'
            },
            'command_injection': {
                'patterns': [
                    r';[\s]*(ls|cat|wget|curl|bash|sh)[\s]',
                    r'\|[\s]*(ls|cat|wget|curl|bash|sh)[\s]',
                    r'`(ls|cat|wget|curl|bash|sh)`',
                    r'\$\((ls|cat|wget|curl|bash|sh)\)',
                    r'(ping|nc|netcat|ncat)[\s]+[\d\.]+',
                    r'&[\s]*(ls|cat|wget|curl|bash|sh)[\s]',
                    r'{\$.*}',
                    r'(chmod|chown|chgrp)[\s]+[\d]+',
                    r'(kill|killall|pkill)[\s]+',
                    r'(python|perl|ruby|php)[\s]+-[ce]',
                    r'(whoami|id|uname|hostname)',
                    r'(passwd|shadow|sudoers)',
                    r'/bin/(sh|bash|dash|ksh|tcsh|zsh)'
                ],
                'severity': 'CRITICAL'
            },
            'xxe': {
                'patterns': [
                    r'<!DOCTYPE[^>]*\[<!ENTITY',
                    r'SYSTEM[\s]*["\']file://',
                    r'SYSTEM[\s]*["\']http://',
                    r'SYSTEM[\s]*["\']https://',
                    r'<!ENTITY[\s]+\w+[\s]+SYSTEM',
                    r'&xxe;',
                    r'%xxe;'
                ],
                'severity': 'HIGH'
            },
            'ldap_injection': {
                'patterns': [
                    r'\*\|\(\w+=\*\)',
                    r'\(\|\(\w+=\*\)\)',
                    r'\(\&\(\w+=\*\)\)',
                    r'[\*\(\)\|&=]'
                ],
                'severity': 'MEDIUM'
            },
            'xpath_injection': {
                'patterns': [
                    r'[\'\"][\s]*or[\s]*[\'\"]',
                    r'[\'\"][\s]*and[\s]*[\'\"]',
                    r'[/\*]+',
                    r'count\(\/\/',
                    r'name\(\)'
                ],
                'severity': 'MEDIUM'
            },
            'ssti': {
                'patterns': [
                    r'{{.*}}',
                    r'{%.*%}',
                    r'\${.*}',
                    r'<%.*%>',
                    r'#\{.*\}',
                    r'{{[\s]*config[\s]*}}',
                    r'{{[\s]*request[\s]*}}',
                    r'{{[\s]*self[\s]*}}'
                ],
                'severity': 'HIGH'
            },
            'log4j': {
                'patterns': [
                    r'\$\{jndi:',
                    r'\$\{.*:.*:.*\}',
                    r'jndi:ldap://',
                    r'jndi:rmi://',
                    r'jndi:dns://'
                ],
                'severity': 'CRITICAL'
            },
            'wordpress_specific': {
                'patterns': [
                    r'/wp-admin/admin-ajax\.php.*action=.*',
                    r'/xmlrpc\.php',
                    r'/wp-json/wp/v2/users',
                    r'/wp-content/plugins/.*\.php\?',
                    r'/wp-content/themes/.*\.php\?',
                    r'/wp-content/uploads/.*\.php',
                    r'/wp-config\.php',
                    r'/wp-login\.php.*log=.*pwd=',
                    r'author=\d+',
                    r'/\?author=\d+'
                ],
                'severity': 'MEDIUM'
            },
            'opencart_specific': {
                'patterns': [
                    r'/admin/controller/.*\.php',
                    r'/system/storage/.*',
                    r'/system/library/.*\.php\?',
                    r'/catalog/controller/.*\.php\?',
                    r'/admin/config\.php',
                    r'/index\.php\?route=.*[\'"]',
                    r'/journal2/.*',
                    r'/vqmod/.*\.php'
                ],
                'severity': 'MEDIUM'
            },
            'scanner_detection': {
                'patterns': [
                    r'/\.git/',
                    r'/\.svn/',
                    r'/\.env',
                    r'/\.aws/',
                    r'/\.docker/',
                    r'/backup\.',
                    r'/database\.',
                    r'/db\.',
                    r'/dump\.',
                    r'/(test|demo|dev|staging|backup|old|temp|tmp)/',
                    r'/phpinfo\.php',
                    r'/info\.php',
                    r'/test\.php',
                    r'/robots\.txt',
                    r'/sitemap\.xml',
                    r'/crossdomain\.xml',
                    r'/clientaccesspolicy\.xml',
                    r'/\.well-known/',
                    r'/admin\.php',
                    r'/login\.php',
                    r'/shell\.php',
                    r'/c99\.php',
                    r'/r57\.php',
                    r'/backdoor\.php'
                    r'/upload\.php'
                    r'/cmd\.php'
                ],
                'severity': 'MEDIUM'
            },
            'authentication_bypass': {
                'patterns': [
                    r'admin[\'"\s]*=[\'"\s]*[\'"]',
                    r'[\'"][\s]*or[\s]*1[\s]*=[\s]*1',
                    r'[\'"][\s]*or[\s]*[\'"]1[\'"][\s]*=[\'"]1',
                    r'admin\'--',
                    r'admin"--',
                    r'admin\#',
                    r'[\'\"][\s]*or[\s]*true',
                    r'[\'\"][\s]*or[\s]*[\'\"]x[\'\"][\s]*=[\s]*[\'\"]x'
                ],
                'severity': 'CRITICAL'
            },
            'directory_listing': {
                'patterns': [
                    r'/\?C=[NMSD];O=[AD]',
                    r'/\?dir',
                    r'/\?sort=',
                    r'/%3f(C|O|M|D|N|S)=',
                    r'/index\.php\?path=',
                    r'/index\.php\?dir='
                ],
                'severity': 'LOW'
            },
            'sensitive_files': {
                'patterns': [
                    r'\.(sql|bak|backup|old|orig|original|copy|tmp|temp|log|swp)($|\?)',
                    r'/(id_rsa|id_dsa|id_ecdsa|id_ed25519)($|\.pub)',
                    r'/\.(ssh|bash_history|zsh_history|mysql_history)/',
                    r'/wp-config\.php\.(bak|old|backup|swp)',
                    r'/(config|configuration|database|db)\.(inc|ini|conf|cfg|json|yml|yaml)',
                    r'/\.(htaccess|htpasswd)($|\.)',
                    r'/web\.config($|\.)',
                    r'/(composer|package)\.(json|lock)',
                    r'/Dockerfile',
                    r'/docker-compose\.yml',
                    r'/\.env($|\.)',
                    r'/(private|secret|key|pass|passwd|password)[\w]*\.'
                ],
                'severity': 'HIGH'
            }
        }
        
        # User-Agent های مشکوک
        self.suspicious_user_agents = {
            'hacking_tools': [
                'nikto', 'sqlmap', 'nmap', 'masscan', 'metasploit',
                'burpsuite', 'burp', 'zaproxy', 'zap', 'acunetix',
                'nessus', 'qualys', 'openvas', 'nexpose', 'appscan',
                'wpscan', 'joomscan', 'droopescan', 'cmsmap',
                'dirbuster', 'dirb', 'gobuster', 'wfuzz', 'ffuf',
                'hydra', 'medusa', 'brutus', 'john', 'hashcat',
                'havij', 'sqlninja', 'pangolin', 'absinthe',
                'beef', 'commix', 'shellshock', 'struct',
                'nuclei', 'jaeles', 'xray', 'rad', 'crawlergo',
                'subfinder', 'httpx', 'naabu', 'meg',
                'python-requests/', 'python-urllib/', 'go-http-client',
                'curl/', 'wget/', 'libwww-perl', 'lwp-trivial'
            ],
            'suspicious_bots': [
                'bot', 'crawler', 'spider', 'scraper', 'scan',
                'fetch', 'check', 'monitor', 'test', 'debug',
                'validator', 'analyzer', 'extractor'
            ],
            'programming_libraries': [
                'python', 'perl', 'ruby', 'java', 'php',
                'node', 'golang', 'rust', 'lua',
                'mechanize', 'scrapy', 'beautifulsoup',
                'httpclient', 'okhttp', 'retrofit'
            ]
        }
        

        # HTTP Methods
        self.suspicious_methods = ['PUT', 'DELETE', 'TRACE', 'CONNECT', 'PATCH', 'OPTIONS']
        
        # Response codes analysis
        self.error_codes = {
            '400': 'Bad Request',
            '401': 'Unauthorized',
            '403': 'Forbidden',
            '404': 'Not Found',
            '405': 'Method Not Allowed',
            '500': 'Internal Server Error',
            '502': 'Bad Gateway',
            '503': 'Service Unavailable'
        }
    
    def extract_file(self) -> List[str]:
        """استخراج فایل gz/zip/tar.gz و برگرداندن لیست فایل‌های لاگ"""
        file_path = Path(self.log_file_path)
        extracted_files = []

        # ایجاد پوشه موقت برای استخراج
        extract_dir = Path('extracted_logs')
        extract_dir.mkdir(exist_ok=True)

        # پردازش tar.gz
        if file_path.suffix == '.gz' and file_path.name.endswith('.tar.gz'):
            print(f"📦 استخراج آرشیو tar.gz: {file_path.name}...")

            with tarfile.open(file_path, 'r:gz') as tar:
                # لیست فایل‌های درون آرشیو
                members = tar.getmembers()
                log_members = []

                # فیلتر کردن فایل‌های لاگ
                for member in members:
                    if member.isfile():
                        name_lower = member.name.lower()
                        if 'access' in name_lower or 'log' in name_lower:
                            log_members.append(member)
                            print(f"  📄 یافت شد: {member.name}")

                # استخراج فایل‌های لاگ
                for member in log_members:
                    tar.extract(member, extract_dir)
                    extracted_path = extract_dir / member.name

                    # اگر فایل استخراج شده هم فشرده است
                    if extracted_path.suffix == '.gz':
                        print(f"    📦 استخراج فایل فشرده: {extracted_path.name}...")
                        decompressed_path = extracted_path.with_suffix('')

                        with gzip.open(extracted_path, 'rb') as gz_file:
                            with open(decompressed_path, 'wb') as out_file:
                                out_file.write(gz_file.read())

                        extracted_files.append(str(decompressed_path))
                        # حذف فایل فشرده موقت
                        extracted_path.unlink()
                    else:
                        extracted_files.append(str(extracted_path))

            print(f"✅ تعداد {len(extracted_files)} فایل لاگ استخراج شد")

        # پردازش فایل gz معمولی
        elif file_path.suffix == '.gz':
            print(f"📦 استخراج {file_path.name}...")
            extracted_path = file_path.with_suffix('')

            with gzip.open(file_path, 'rb') as gz_file:
                with open(extracted_path, 'wb') as output_file:
                    content = gz_file.read()
                    output_file.write(content)

            extracted_files.append(str(extracted_path))
            print(f"✅ استخراج شد: {extracted_path.name}")

        # پردازش فایل zip
        elif file_path.suffix == '.zip':
            print(f"📦 استخراج {file_path.name}...")

            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                # لیست فایل‌های درون zip
                for file_info in zip_ref.filelist:
                    name_lower = file_info.filename.lower()
                    if 'access' in name_lower or 'log' in name_lower:
                        print(f"  📄 استخراج: {file_info.filename}")
                        zip_ref.extract(file_info, extract_dir)
                        extracted_path = extract_dir / file_info.filename

                        # اگر فایل استخراج شده هم فشرده است
                        if extracted_path.suffix == '.gz':
                            print(f"    📦 استخراج فایل فشرده: {extracted_path.name}...")
                            decompressed_path = extracted_path.with_suffix('')

                            with gzip.open(extracted_path, 'rb') as gz_file:
                                with open(decompressed_path, 'wb') as out_file:
                                    out_file.write(gz_file.read())

                            extracted_files.append(str(decompressed_path))
                            extracted_path.unlink()
                        else:
                            extracted_files.append(str(extracted_path))

            print(f"✅ تعداد {len(extracted_files)} فایل لاگ استخراج شد")

        # فایل معمولی
        else:
            extracted_files.append(str(file_path))

        return extracted_files if extracted_files else [str(file_path)]

    def parse_log_line(self, line: str) -> Dict:
        """پارس خط لاگ با پشتیبانی از فرمت‌های مختلف"""
        # Combined Log Format
        pattern = r'(\S+) - - \[(.*?)\] "(.*?)" (\d+) (\d+|-) "(.*?)" "(.*?)"'
        match = re.match(pattern, line)
        
        if match:
            # پارس HTTP method و URL
            request = match.group(3)
            request_parts = request.split(' ')
            method = request_parts[0] if len(request_parts) > 0 else 'UNKNOWN'
            url = request_parts[1] if len(request_parts) > 1 else '/'
            protocol = request_parts[2] if len(request_parts) > 2 else 'HTTP/1.0'
            
            # پارس timestamp
            timestamp_str = match.group(2)
            try:
                # Format: "21/Sep/2025:15:42:03 +0330"
                dt = datetime.strptime(timestamp_str.split(' ')[0], '%d/%b/%Y:%H:%M:%S')
            except:
                dt = datetime.now()
            
            return {
                'ip': match.group(1),
                'timestamp': timestamp_str,
                'datetime': dt,
                'request': request,
                'method': method,
                'url': url,
                'protocol': protocol,
                'status_code': int(match.group(4)),
                'bytes': int(match.group(5)) if match.group(5) != '-' else 0,
                'referrer': match.group(6),
                'user_agent': match.group(7)
            }
        
        return None

    def select_time_period(self) -> int:
        """انتخاب بازه زمانی برای تحلیل"""
        print("\n" + "="*60)
        print("📅 انتخاب بازه زمانی برای تحلیل:")
        print("-"*60)
        print("1. یک ماه اخیر")
        print("2. دو ماه اخیر") 
        print("3. سه ماه اخیر")
        print("4. شش ماه اخیر")
        print("5. دوازده ماه اخیر")
        print("6. کل لاگ‌ها (بدون محدودیت زمانی)")
        print("-"*60)

        while True:
            try:
                choice = input("🔢 گزینه مورد نظر را انتخاب کنید (1-6): ").strip()
                if choice in ['1', '2', '3', '4', '5', '6']:
                    # تبدیل به تعداد روز
                    days_map = {
                        '1': 30,    # 1 ماه
                        '2': 60,    # 2 ماه
                        '3': 90,    # 3 ماه
                        '4': 180,   # 6 ماه
                        '5': 365,   # 12 ماه
                        '6': 0      # کل (بدون محدودیت)
                    }

                    period_names = {
                        '1': 'یک ماه اخیر',
                        '2': 'دو ماه اخیر',
                        '3': 'سه ماه اخیر',
                        '4': 'شش ماه اخیر',
                        '5': 'دوازده ماه اخیر',
                        '6': 'کل لاگ‌ها'
                    }

                    print(f"\n✅ بازه انتخاب شده: {period_names[choice]}")
                    return days_map[choice]
                else:
                    print("❌ لطفاً عددی بین 1 تا 6 وارد کنید")
            except KeyboardInterrupt:
                print("\n⚠️ عملیات لغو شد")
                sys.exit(0)
            except Exception as e:
                print(f"❌ خطا: {e}")

    def load_logs(self, days_limit: int = 0) -> List[Dict]:
        """بارگذاری و پارس لاگ‌ها با پشتیبانی از چندین فایل"""
        log_files = self.extract_file()

        print(f"📖 خواندن {len(log_files)} فایل لاگ...")

        # محاسبه تاریخ cutoff
        cutoff_date = None
        if days_limit > 0:
            cutoff_date = datetime.now() - timedelta(days=days_limit)
            print(f"🕐 فیلتر زمانی: لاگ‌های بعد از {cutoff_date.strftime('%Y-%m-%d')}")

        total_lines = 0
        parsed_lines = 0
        filtered_lines = 0
        file_stats = {}

        # پردازش هر فایل لاگ
        for log_file in sorted(log_files):
            print(f"\n  📄 پردازش: {Path(log_file).name}")
            file_total = 0
            file_parsed = 0
            file_filtered = 0

            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']

            for encoding in encodings:
                try:
                    with open(log_file, 'r', encoding=encoding, errors='ignore') as f:
                        for line_num, line in enumerate(f, 1):
                            file_total += 1
                            total_lines += 1

                            if total_lines % 10000 == 0:
                                print(f"    پردازش: {total_lines:,} خط...")

                            parsed = self.parse_log_line(line.strip())
                            if parsed:
                                file_parsed += 1
                                parsed_lines += 1

                                # بررسی محدودیت زمانی
                                if cutoff_date:
                                    if parsed['datetime'] < cutoff_date:
                                        file_filtered += 1
                                        filtered_lines += 1
                                        continue
                                    
                                # اضافه کردن نام فایل منبع
                                parsed['source_file'] = Path(log_file).name
                                self.logs.append(parsed)

                    file_stats[Path(log_file).name] = {
                        'total': file_total,
                        'parsed': file_parsed,
                        'filtered': file_filtered,
                        'loaded': file_parsed - file_filtered
                    }

                    print(f"    ✓ {Path(log_file).name}:")
                    print(f"      کل خطوط: {file_total:,}")
                    print(f"      پارس شده: {file_parsed:,}")
                    if days_limit > 0:
                        print(f"      فیلتر شده: {file_filtered:,}")
                        print(f"      بارگذاری شده: {file_parsed - file_filtered:,}")

                    break

                except UnicodeDecodeError:
                    if encoding == encodings[-1]:
                        print(f"    ❌ خطا در خواندن {Path(log_file).name}")
                    continue
                
        # مرتب‌سازی لاگ‌ها بر اساس زمان
        self.logs.sort(key=lambda x: x['datetime'])

        # نمایش خلاصه نهایی
        print(f"\n✅ بارگذاری کامل:")
        print(f"  • تعداد فایل‌ها: {len(log_files)}")
        print(f"  • کل خطوط: {total_lines:,}")
        print(f"  • پارس شده: {parsed_lines:,}")
        if days_limit > 0:
            print(f"  • فیلتر شده (قدیمی): {filtered_lines:,}")
            print(f"  • در بازه زمانی: {len(self.logs):,}")
        else:
            print(f"  • بارگذاری شده: {len(self.logs):,}")
        print(f"  • نادیده گرفته: {total_lines - parsed_lines:,}")

        # نمایش بازه زمانی واقعی
        if self.logs:
            date_range = {
                'start': min(log['datetime'] for log in self.logs),
                'end': max(log['datetime'] for log in self.logs)
            }
            print(f"  • بازه زمانی: {date_range['start'].strftime('%Y-%m-%d')} تا {date_range['end'].strftime('%Y-%m-%d')}")

        # نمایش آمار هر فایل
        if len(file_stats) > 1:
            print(f"\n📊 آمار فایل‌ها:")
            for filename, stats in file_stats.items():
                print(f"  • {filename}: {stats['loaded']:,} لاگ")

        # حذف فایل‌های موقت
        self._cleanup_temp_files()

        return self.logs
    
    def _cleanup_temp_files(self):
        """پاکسازی فایل‌های موقت استخراج شده"""
        extract_dir = Path('extracted_logs')
        if extract_dir.exists():
            try:
                import shutil
                shutil.rmtree(extract_dir)
                print("🧹 فایل‌های موقت پاکسازی شدند")
            except Exception as e:
                print(f"⚠️ خطا در پاکسازی فایل‌های موقت: {e}")

    def calculate_ip_risk_score(self, ip: str) -> Dict:
        """محاسبه امتیاز ریسک برای هر IP"""
        score = 0
        reasons = []
        
        # جمع‌آوری داده‌های IP
        ip_logs = [log for log in self.logs if log['ip'] == ip]
        
        # 1. تعداد درخواست‌ها
        request_count = len(ip_logs)
        if request_count > 1000:
            score += 30
            reasons.append(f"تعداد درخواست زیاد ({request_count})")
        elif request_count > 500:
            score += 20
            reasons.append(f"تعداد درخواست متوسط ({request_count})")
        elif request_count > 100:
            score += 10
            reasons.append(f"تعداد درخواست قابل توجه ({request_count})")
        
        # 2. تحلیل User-Agent
        user_agents = set(log['user_agent'] for log in ip_logs)
        suspicious_ua_count = 0
        
        for ua in user_agents:
            ua_lower = ua.lower()
            # ابزارهای هک
            for tool in self.suspicious_user_agents['hacking_tools']:
                if tool in ua_lower:
                    score += 25
                    suspicious_ua_count += 1
                    reasons.append(f"ابزار هک: {tool}")
                    break
            
            # User-Agent خالی
            if ua == '-' or len(ua) < 5:
                score += 15
                reasons.append("User-Agent خالی/کوتاه")
        
        # 3. HTTP Methods مشکوک
        methods = Counter(log['method'] for log in ip_logs)
        for method in self.suspicious_methods:
            if method in methods:
                score += 10
                reasons.append(f"HTTP Method مشکوک: {method}")
        
        # 4. کدهای خطا
        status_codes = Counter(log['status_code'] for log in ip_logs)
        error_count = sum(count for code, count in status_codes.items() if code >= 400)
        if error_count > 50:
            score += 20
            reasons.append(f"تعداد خطای زیاد ({error_count})")
        elif error_count > 20:
            score += 10
            reasons.append(f"تعداد خطای متوسط ({error_count})")
        
        # 5. بررسی الگوهای حمله
        attack_count = 0
        attack_types = set()
        
        for log in ip_logs:
            request = log['request']
            for attack_type, attack_info in self.advanced_attack_patterns.items():
                for pattern in attack_info['patterns']:
                    if re.search(pattern, request, re.IGNORECASE):
                        attack_count += 1
                        attack_types.add(attack_type)
                        
                        if attack_info['severity'] == 'CRITICAL':
                            score += 30
                        elif attack_info['severity'] == 'HIGH':
                            score += 20
                        elif attack_info['severity'] == 'MEDIUM':
                            score += 10
                        else:
                            score += 5
                        break
        
        if attack_count > 0:
            reasons.append(f"تعداد حملات: {attack_count} ({', '.join(attack_types)})")
        
        # 6. نرخ درخواست (requests per minute)
        if ip_logs:
            timestamps = sorted([log['datetime'] for log in ip_logs])
            time_range = (timestamps[-1] - timestamps[0]).total_seconds() / 60  # دقیقه
            
            if time_range > 0:
                requests_per_minute = len(ip_logs) / time_range
                if requests_per_minute > 100:
                    score += 30
                    reasons.append(f"نرخ درخواست بالا ({requests_per_minute:.1f}/min)")
                elif requests_per_minute > 50:
                    score += 20
                    reasons.append(f"نرخ درخواست متوسط ({requests_per_minute:.1f}/min)")
        
        # 7. تنوع URL ها (برای تشخیص اسکنرها)
        unique_urls = len(set(log['url'] for log in ip_logs))
        if unique_urls > 100:
            score += 20
            reasons.append(f"اسکن URL ({unique_urls} URL متفاوت)")
        elif unique_urls > 50:
            score += 10
            reasons.append(f"تنوع URL بالا ({unique_urls} URL)")
        
        # 8. دسترسی به فایل‌های حساس
        sensitive_access = 0
        for log in ip_logs:
            for pattern in self.advanced_attack_patterns['sensitive_files']['patterns']:
                if re.search(pattern, log['url'], re.IGNORECASE):
                    sensitive_access += 1
                    score += 10
                    break
        
        if sensitive_access > 0:
            reasons.append(f"دسترسی به فایل‌های حساس ({sensitive_access} مورد)")
        
        # تعیین سطح ریسک
        if score >= 100:
            risk_level = 'CRITICAL'
        elif score >= 70:
            risk_level = 'HIGH'
        elif score >= 40:
            risk_level = 'MEDIUM'
        elif score >= 20:
            risk_level = 'LOW'
        else:
            risk_level = 'SAFE'
        
        return {
            'ip': ip,
            'score': score,
            'risk_level': risk_level,
            'reasons': reasons,
            'request_count': request_count,
            'unique_urls': unique_urls,
            'error_count': error_count
        }
    
    def analyze_temporal_patterns(self) -> Dict:
        """تحلیل الگوهای زمانی حملات"""
        hourly_stats = defaultdict(int)
        daily_stats = defaultdict(int)
        
        for log in self.logs:
            dt = log['datetime']
            hourly_stats[dt.hour] += 1
            daily_stats[dt.strftime('%Y-%m-%d')] += 1
        
        # شناسایی ساعات پرترافیک
        peak_hours = sorted(hourly_stats.items(), key=lambda x: x[1], reverse=True)[:5]
        
        # شناسایی روزهای پرحمله
        peak_days = sorted(daily_stats.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return {
            'hourly_distribution': dict(hourly_stats),
            'daily_distribution': dict(daily_stats),
            'peak_hours': peak_hours,
            'peak_days': peak_days
        }
    
    def analyze_geographic_patterns(self) -> Dict:
        """تحلیل جغرافیایی IP ها (نیاز به API دارد)"""
        # این بخش می‌تواند با استفاده از سرویس‌های GeoIP پیاده‌سازی شود
        geo_stats = {
            'countries': defaultdict(int),
            'suspicious_countries': []
        }
        
        # لیست کشورهای پرخطر (مثال)
        high_risk_countries = ['CN', 'RU', 'KP', 'IR']
        
        # برای نمونه، از IP ranges ساده استفاده می‌کنیم
        for log in self.logs:
            ip = log['ip']
            # این قسمت نیاز به پیاده‌سازی واقعی GeoIP دارد
            # geo_stats['countries']['Unknown'] += 1
        
        return geo_stats
    
    def analyze_bots(self) -> Dict:
        """تحلیل سریع بات‌ها با Threading و بدون DNS"""
        import concurrent.futures
        from threading import Lock
        import time
        from collections import defaultdict, Counter

        print("    ⚡ استفاده از Threading برای تحلیل سریع بات‌ها (بدون DNS)...")
        start_time = time.time()

        # ساختار نتایج با thread-safe locks
        result_lock = Lock()

        bot_analysis = {
            'legitimate': defaultdict(lambda: {
                'ips': set(), 
                'requests': 0, 
                'unique_urls': set(), 
                'first_seen': None, 
                'last_seen': None,
                'ip_requests': defaultdict(int)  # اضافه شده: شمارش درخواست هر IP
            }),
            'potentially_legitimate': defaultdict(lambda: {
                'ips': set(), 
                'requests': 0, 
                'unique_urls': set(), 
                'first_seen': None, 
                'last_seen': None,
                'ip_requests': defaultdict(int)  # اضافه شده
            }),
            'fake': defaultdict(lambda: {
                'ips': set(), 
                'requests': 0, 
                'user_agents': Counter(), 
                'patterns': Counter(), 
                'first_seen': None, 
                'last_seen': None,
                'ip_requests': defaultdict(int)  # اضافه شده
            }),
            'unknown': {
                'ips': set(), 
                'requests': 0, 
                'user_agents': Counter(), 
                'unique_urls': set(), 
                'first_seen': None, 
                'last_seen': None,
                'ip_requests': defaultdict(int)  # اضافه شده
            },
            'bot_activity': defaultdict(lambda: defaultdict(int)),  # مطمئن شدن از وجود
            'bot_traffic_distribution': defaultdict(lambda: defaultdict(int)),
            'bot_ip_distribution': defaultdict(lambda: defaultdict(int))
        }

        # Progress tracking
        processed_count = 0
        progress_lock = Lock()

        def process_log_chunk(logs_chunk, chunk_id):
            """پردازش یک بخش از لاگ‌ها بدون DNS"""
            local_results = {
                'legitimate': defaultdict(lambda: {
                    'ips': set(), 
                    'requests': 0, 
                    'unique_urls': set(), 
                    'first_seen': None, 
                    'last_seen': None,
                    'ip_requests': defaultdict(int)  # اضافه شده
                }),
                'potentially_legitimate': defaultdict(lambda: {
                    'ips': set(), 
                    'requests': 0, 
                    'unique_urls': set(), 
                    'first_seen': None, 
                    'last_seen': None,
                    'ip_requests': defaultdict(int)  # اضافه شده
                }),
                'fake': defaultdict(lambda: {
                    'ips': set(), 
                    'requests': 0, 
                    'user_agents': Counter(), 
                    'patterns': Counter(), 
                    'first_seen': None, 
                    'last_seen': None,
                    'ip_requests': defaultdict(int)  # اضافه شده
                }),
                'unknown': {
                    'ips': set(), 
                    'requests': 0, 
                    'user_agents': Counter(), 
                    'unique_urls': set(), 
                    'first_seen': None, 
                    'last_seen': None,
                    'ip_requests': defaultdict(int)  # اضافه شده
                },
                'bot_activity': defaultdict(lambda: defaultdict(int)),
                'bot_traffic_distribution': defaultdict(lambda: defaultdict(int))
            }

            # کش محلی برای IP ها
            local_ip_cache = {}

            for log in logs_chunk:
                ip = log['ip']
                ua = log['user_agent']
                url = log['url']
                dt = log['datetime']
                ua_lower = ua.lower()

                # کلید کش
                cache_key = f"{ip}:{ua}"

                # چک کش محلی
                if cache_key in local_ip_cache:
                    bot_type, bot_category = local_ip_cache[cache_key]

                    if bot_category == 'legitimate':
                        local_results['legitimate'][bot_type]['ips'].add(ip)
                        local_results['legitimate'][bot_type]['requests'] += 1
                        local_results['legitimate'][bot_type]['ip_requests'][ip] += 1  # اضافه شده
                        local_results['legitimate'][bot_type]['unique_urls'].add(url)
                        if not local_results['legitimate'][bot_type]['first_seen'] or dt < local_results['legitimate'][bot_type]['first_seen']:
                            local_results['legitimate'][bot_type]['first_seen'] = dt
                        if not local_results['legitimate'][bot_type]['last_seen'] or dt > local_results['legitimate'][bot_type]['last_seen']:
                            local_results['legitimate'][bot_type]['last_seen'] = dt
                        local_results['bot_activity'][url][bot_type] += 1
                        local_results['bot_traffic_distribution'][bot_type][dt.hour] += 1

                    elif bot_category == 'potentially_legitimate':
                        local_results['potentially_legitimate'][bot_type]['ips'].add(ip)
                        local_results['potentially_legitimate'][bot_type]['requests'] += 1
                        local_results['potentially_legitimate'][bot_type]['ip_requests'][ip] += 1  # اضافه شده
                        local_results['potentially_legitimate'][bot_type]['unique_urls'].add(url)
                        local_results['bot_activity'][url][bot_type] += 1
                        local_results['bot_traffic_distribution'][bot_type][dt.hour] += 1

                    elif bot_category == 'fake':
                        local_results['fake'][bot_type]['ips'].add(ip)
                        local_results['fake'][bot_type]['requests'] += 1
                        local_results['fake'][bot_type]['ip_requests'][ip] += 1  # اضافه شده
                        local_results['fake'][bot_type]['user_agents'][ua] += 1
                        local_results['bot_activity'][url][bot_type] += 1
                        local_results['bot_traffic_distribution'][bot_type][dt.hour] += 1

                    else:  # unknown
                        local_results['unknown']['ips'].add(ip)
                        local_results['unknown']['requests'] += 1
                        local_results['unknown']['ip_requests'][ip] += 1  # اضافه شده
                        local_results['unknown']['unique_urls'].add(url)
                        local_results['bot_activity'][url]['Unknown'] += 1
                        local_results['bot_traffic_distribution']['Unknown'][dt.hour] += 1

                    continue
                
                # ... (بقیه کد تشخیص بات که قبلاً بود)

                # تشخیص نوع بات بدون DNS
                identified = False
                bot_type = None
                bot_category = None

                # بررسی بات‌های معتبر
                for bot_name, bot_info in self.legitimate_bots.items():
                    if not bot_info['patterns']:
                        continue

                    # بررسی User-Agent
                    if any(pattern in ua_lower for pattern in bot_info['patterns']):
                        # بررسی IP range اگر موجود باشد
                        ip_matches = False
                        if bot_info['ip_ranges']:
                            try:
                                ip_obj = ipaddress.ip_address(ip)
                                # فقط 5 range اول را چک کن برای سرعت
                                for ip_range in bot_info['ip_ranges'][:5]:
                                    try:
                                        if '/' in ip_range:
                                            network = ipaddress.ip_network(ip_range, strict=False)
                                            if ip_obj in network:
                                                ip_matches = True
                                                break
                                    except:
                                        continue
                                    
                                if ip_matches:
                                    bot_type = bot_name
                                    bot_category = 'legitimate'
                                else:
                                    # UA درست، IP غلط
                                    bot_type = bot_name
                                    bot_category = 'potentially_legitimate'
                            except:
                                # در صورت خطا، فقط بر اساس UA
                                bot_type = bot_name
                                bot_category = 'legitimate'
                        else:
                            # اگر IP range نداریم، فقط بر اساس UA
                            bot_type = bot_name
                            bot_category = 'legitimate'

                        identified = True
                        break
                    
                # بررسی بات‌های جعلی
                if not identified:
                    # ابزارهای هک
                    for tool in ['nikto', 'sqlmap', 'nmap', 'burp', 'acunetix', 'wpscan', 'metasploit', 
                                'python-requests', 'curl/', 'wget/', 'libwww-perl', 'python/', 'scrapy']:
                        if tool in ua_lower:
                            bot_type = f"Hacking Tool: {tool}"
                            bot_category = 'fake'
                            identified = True
                            break
                        
                    # User-Agent خالی یا مشکوک
                    if not identified:
                        if ua == '-' or len(ua) < 5:
                            bot_type = "Empty/Invalid UA"
                            bot_category = 'fake'
                            identified = True
                        elif any(word in ua_lower for word in ['bot', 'crawler', 'spider', 'scraper']):
                            # بات‌های مشکوک که در لیست معتبر نیستند
                            for legit_bot in self.legitimate_bots.keys():
                                if legit_bot.lower() in ua_lower:
                                    break
                            else:
                                bot_type = "Suspicious Bot"
                                bot_category = 'fake'
                                identified = True

                # اگر شناسایی نشد، ناشناس است
                if not identified:
                    bot_type = 'Unknown'
                    bot_category = 'unknown'

                # ذخیره در کش محلی
                local_ip_cache[cache_key] = (bot_type, bot_category)

                # ذخیره داده‌ها بر اساس دسته‌بندی
                if bot_category == 'legitimate':
                    local_results['legitimate'][bot_type]['ips'].add(ip)
                    local_results['legitimate'][bot_type]['requests'] += 1
                    local_results['legitimate'][bot_type]['ip_requests'][ip] += 1  # اضافه شده
                    local_results['legitimate'][bot_type]['unique_urls'].add(url)
                    if not local_results['legitimate'][bot_type]['first_seen'] or dt < local_results['legitimate'][bot_type]['first_seen']:
                        local_results['legitimate'][bot_type]['first_seen'] = dt
                    if not local_results['legitimate'][bot_type]['last_seen'] or dt > local_results['legitimate'][bot_type]['last_seen']:
                        local_results['legitimate'][bot_type]['last_seen'] = dt
                    local_results['bot_activity'][url][bot_type] += 1
                    local_results['bot_traffic_distribution'][bot_type][dt.hour] += 1

                elif bot_category == 'potentially_legitimate':
                    local_results['potentially_legitimate'][bot_type]['ips'].add(ip)
                    local_results['potentially_legitimate'][bot_type]['requests'] += 1
                    local_results['potentially_legitimate'][bot_type]['ip_requests'][ip] += 1  # اضافه شده
                    local_results['potentially_legitimate'][bot_type]['unique_urls'].add(url)
                    if not local_results['potentially_legitimate'][bot_type]['first_seen'] or dt < local_results['potentially_legitimate'][bot_type]['first_seen']:
                        local_results['potentially_legitimate'][bot_type]['first_seen'] = dt
                    if not local_results['potentially_legitimate'][bot_type]['last_seen'] or dt > local_results['potentially_legitimate'][bot_type]['last_seen']:
                        local_results['potentially_legitimate'][bot_type]['last_seen'] = dt
                    local_results['bot_activity'][url][bot_type] += 1
                    local_results['bot_traffic_distribution'][bot_type][dt.hour] += 1

                elif bot_category == 'fake':
                    local_results['fake'][bot_type]['ips'].add(ip)
                    local_results['fake'][bot_type]['requests'] += 1
                    local_results['fake'][bot_type]['ip_requests'][ip] += 1  # اضافه شده
                    local_results['fake'][bot_type]['user_agents'][ua] += 1
                    local_results['fake'][bot_type]['patterns'][ua_lower] += 1
                    if not local_results['fake'][bot_type]['first_seen'] or dt < local_results['fake'][bot_type]['first_seen']:
                        local_results['fake'][bot_type]['first_seen'] = dt
                    if not local_results['fake'][bot_type]['last_seen'] or dt > local_results['fake'][bot_type]['last_seen']:
                        local_results['fake'][bot_type]['last_seen'] = dt
                    local_results['bot_activity'][url][bot_type] += 1
                    local_results['bot_traffic_distribution'][bot_type][dt.hour] += 1

                else:  # unknown
                    local_results['unknown']['ips'].add(ip)
                    local_results['unknown']['requests'] += 1
                    local_results['unknown']['ip_requests'][ip] += 1  # اضافه شده
                    local_results['unknown']['user_agents'][ua] += 1
                    local_results['unknown']['unique_urls'].add(url)
                    if not local_results['unknown']['first_seen'] or dt < local_results['unknown']['first_seen']:
                        local_results['unknown']['first_seen'] = dt
                    if not local_results['unknown']['last_seen'] or dt > local_results['unknown']['last_seen']:
                        local_results['unknown']['last_seen'] = dt
                    local_results['bot_activity'][url]['Unknown'] += 1
                    local_results['bot_traffic_distribution']['Unknown'][dt.hour] += 1

            # آپدیت progress
            nonlocal processed_count
            with progress_lock:
                processed_count += len(logs_chunk)
                progress = (processed_count / len(self.logs)) * 100
                if chunk_id % 5 == 0:
                    elapsed = time.time() - start_time
                    speed = processed_count / elapsed if elapsed > 0 else 0
                    eta = (len(self.logs) - processed_count) / speed if speed > 0 else 0
                    print(f"      Thread {chunk_id}: {processed_count:,}/{len(self.logs):,} ({progress:.1f}%) - سرعت: {speed:.0f} log/s - ETA: {eta:.0f}s")

            return local_results

        def merge_results(main_results, chunk_results):
            """ادغام نتایج chunk با نتایج اصلی"""
            with result_lock:
                # ادغام legitimate
                for bot_name, data in chunk_results['legitimate'].items():
                    main_results['legitimate'][bot_name]['ips'].update(data['ips'])
                    main_results['legitimate'][bot_name]['requests'] += data['requests']
                    main_results['legitimate'][bot_name]['unique_urls'].update(data['unique_urls'])

                    # ادغام ip_requests
                    for ip, count in data['ip_requests'].items():
                        main_results['legitimate'][bot_name]['ip_requests'][ip] += count

                    if data['first_seen']:
                        if main_results['legitimate'][bot_name]['first_seen'] is None or data['first_seen'] < main_results['legitimate'][bot_name]['first_seen']:
                            main_results['legitimate'][bot_name]['first_seen'] = data['first_seen']

                    if data['last_seen']:
                        if main_results['legitimate'][bot_name]['last_seen'] is None or data['last_seen'] > main_results['legitimate'][bot_name]['last_seen']:
                            main_results['legitimate'][bot_name]['last_seen'] = data['last_seen']

                # ادغام potentially_legitimate
                for bot_name, data in chunk_results['potentially_legitimate'].items():
                    main_results['potentially_legitimate'][bot_name]['ips'].update(data['ips'])
                    main_results['potentially_legitimate'][bot_name]['requests'] += data['requests']
                    main_results['potentially_legitimate'][bot_name]['unique_urls'].update(data['unique_urls'])

                    # ادغام ip_requests
                    for ip, count in data['ip_requests'].items():
                        main_results['potentially_legitimate'][bot_name]['ip_requests'][ip] += count

                    if data.get('first_seen'):
                        if main_results['potentially_legitimate'][bot_name]['first_seen'] is None or data['first_seen'] < main_results['potentially_legitimate'][bot_name]['first_seen']:
                            main_results['potentially_legitimate'][bot_name]['first_seen'] = data['first_seen']

                    if data.get('last_seen'):
                        if main_results['potentially_legitimate'][bot_name]['last_seen'] is None or data['last_seen'] > main_results['potentially_legitimate'][bot_name]['last_seen']:
                            main_results['potentially_legitimate'][bot_name]['last_seen'] = data['last_seen']

                # ادغام fake
                for fake_type, data in chunk_results['fake'].items():
                    main_results['fake'][fake_type]['ips'].update(data['ips'])
                    main_results['fake'][fake_type]['requests'] += data['requests']

                    # ادغام ip_requests
                    for ip, count in data['ip_requests'].items():
                        main_results['fake'][fake_type]['ip_requests'][ip] += count

                    for ua, count in data['user_agents'].items():
                        main_results['fake'][fake_type]['user_agents'][ua] += count

                    for pattern, count in data['patterns'].items():
                        main_results['fake'][fake_type]['patterns'][pattern] += count

                    if data.get('first_seen'):
                        if main_results['fake'][fake_type]['first_seen'] is None or data['first_seen'] < main_results['fake'][fake_type]['first_seen']:
                            main_results['fake'][fake_type]['first_seen'] = data['first_seen']

                    if data.get('last_seen'):
                        if main_results['fake'][fake_type]['last_seen'] is None or data['last_seen'] > main_results['fake'][fake_type]['last_seen']:
                            main_results['fake'][fake_type]['last_seen'] = data['last_seen']

                # ادغام unknown
                main_results['unknown']['ips'].update(chunk_results['unknown']['ips'])
                main_results['unknown']['requests'] += chunk_results['unknown']['requests']
                main_results['unknown']['unique_urls'].update(chunk_results['unknown']['unique_urls'])

                # ادغام ip_requests
                for ip, count in chunk_results['unknown']['ip_requests'].items():
                    main_results['unknown']['ip_requests'][ip] += count

                for ua, count in chunk_results['unknown']['user_agents'].items():
                    main_results['unknown']['user_agents'][ua] += count

                if chunk_results['unknown'].get('first_seen'):
                    if main_results['unknown']['first_seen'] is None or chunk_results['unknown']['first_seen'] < main_results['unknown']['first_seen']:
                        main_results['unknown']['first_seen'] = chunk_results['unknown']['first_seen']

                if chunk_results['unknown'].get('last_seen'):
                    if main_results['unknown']['last_seen'] is None or chunk_results['unknown']['last_seen'] > main_results['unknown']['last_seen']:
                        main_results['unknown']['last_seen'] = chunk_results['unknown']['last_seen']

                # ادغام bot_activity
                for url, bots_dict in chunk_results['bot_activity'].items():
                    for bot_type, count in bots_dict.items():
                        main_results['bot_activity'][url][bot_type] += count

                # ادغام bot_traffic_distribution
                for bot_type, hours_dict in chunk_results['bot_traffic_distribution'].items():
                    for hour, count in hours_dict.items():
                        main_results['bot_traffic_distribution'][bot_type][hour] += count

        # تقسیم لاگ‌ها به chunk ها
        total_logs = len(self.logs)
        chunk_size = 3000  # اندازه کوچکتر برای سرعت بیشتر
        num_workers = min(16, max(4, (total_logs // chunk_size) + 1))  # بین 4 تا 16 thread

        chunks = [self.logs[i:i+chunk_size] for i in range(0, total_logs, chunk_size)]
        print(f"      تعداد chunks: {len(chunks)}, تعداد workers: {num_workers}")
        print(f"      پردازش {total_logs:,} لاگ...")

        # اجرای موازی با ThreadPoolExecutor
        with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
            # ارسال همه chunk ها برای پردازش
            futures = []
            for i, chunk in enumerate(chunks):
                future = executor.submit(process_log_chunk, chunk, i)
                futures.append(future)

            # دریافت نتایج و ادغام
            completed = 0
            for future in concurrent.futures.as_completed(futures):
                try:
                    chunk_result = future.result()
                    merge_results(bot_analysis, chunk_result)
                    completed += 1

                    if completed % 10 == 0:
                        print(f"      Chunks completed: {completed}/{len(chunks)}")

                except Exception as e:
                    print(f"      ⚠️ خطا در پردازش chunk: {e}")

        # محاسبه آمار نهایی
        print("      📊 محاسبه آمار نهایی...")

        # آمار بات‌های معتبر
        for bot_type in list(bot_analysis['legitimate'].keys()):
            bot_analysis['legitimate'][bot_type]['ips_count'] = len(bot_analysis['legitimate'][bot_type]['ips'])
            bot_analysis['legitimate'][bot_type]['unique_urls_count'] = len(bot_analysis['legitimate'][bot_type]['unique_urls'])

            # محاسبه top URLs
            top_urls = []
            if 'bot_activity' in bot_analysis:  # بررسی وجود کلید
                for url, bots_dict in bot_analysis['bot_activity'].items():
                    if bot_type in bots_dict:
                        top_urls.append((url, bots_dict[bot_type]))
            bot_analysis['legitimate'][bot_type]['top_urls'] = sorted(top_urls, key=lambda x: x[1], reverse=True)[:10]

        # آمار بات‌های احتمالی
        for bot_type in list(bot_analysis['potentially_legitimate'].keys()):
            bot_analysis['potentially_legitimate'][bot_type]['ips_count'] = len(bot_analysis['potentially_legitimate'][bot_type]['ips'])
            bot_analysis['potentially_legitimate'][bot_type]['unique_urls_count'] = len(bot_analysis['potentially_legitimate'][bot_type]['unique_urls'])

        # آمار بات‌های جعلی
        for bot_type in list(bot_analysis['fake'].keys()):
            bot_analysis['fake'][bot_type]['ips_count'] = len(bot_analysis['fake'][bot_type]['ips'])
            bot_analysis['fake'][bot_type]['top_user_agents'] = bot_analysis['fake'][bot_type]['user_agents'].most_common(5)
            bot_analysis['fake'][bot_type]['top_patterns'] = bot_analysis['fake'][bot_type]['patterns'].most_common(5)

            # محاسبه top URLs
            top_urls = []
            if 'bot_activity' in bot_analysis:  # بررسی وجود کلید
                for url, bots_dict in bot_analysis['bot_activity'].items():
                    if bot_type in bots_dict:
                        top_urls.append((url, bots_dict[bot_type]))
            bot_analysis['fake'][bot_type]['top_urls'] = sorted(top_urls, key=lambda x: x[1], reverse=True)[:10]

        # آمار ناشناس
        bot_analysis['unknown']['ips_count'] = len(bot_analysis['unknown']['ips'])
        bot_analysis['unknown']['unique_urls_count'] = len(bot_analysis['unknown']['unique_urls'])
        bot_analysis['unknown']['top_user_agents'] = bot_analysis['unknown']['user_agents'].most_common(10)

        # محاسبه top URLs برای unknown
        unknown_urls = []
        if 'bot_activity' in bot_analysis:  # بررسی وجود کلید
            for url, bots_dict in bot_analysis['bot_activity'].items():
                if 'Unknown' in bots_dict:
                    unknown_urls.append((url, bots_dict['Unknown']))
        bot_analysis['unknown']['top_urls'] = sorted(unknown_urls, key=lambda x: x[1], reverse=True)[:10]

        elapsed_time = time.time() - start_time
        print(f"      ✅ تحلیل بات‌ها کامل شد در {elapsed_time:.1f} ثانیه")

        # نمایش خلاصه آمار
        total_bot_requests = sum(b['requests'] for b in bot_analysis['legitimate'].values())
        total_potentially = sum(b['requests'] for b in bot_analysis['potentially_legitimate'].values())
        total_fake_requests = sum(b['requests'] for b in bot_analysis['fake'].values())
        total_unknown_requests = bot_analysis['unknown']['requests']

        print(f"      📊 خلاصه: معتبر: {total_bot_requests:,} | احتمالی: {total_potentially:,} | جعلی: {total_fake_requests:,} | ناشناس: {total_unknown_requests:,}")

        return bot_analysis

    def comprehensive_analysis(self) -> Dict:
        """تحلیل جامع همه جنبه‌های امنیتی"""
        print("⚙️ شروع تحلیل جامع امنیتی...")
        analysis = {
            'overview': {},
            'risk_scores': {},
            'bot_analysis': {
                'legitimate': defaultdict(list),
                'fake': defaultdict(list),
                'unknown': []
            },
            'attack_analysis': defaultdict(lambda: defaultdict(list)),
            'temporal_analysis': {},
            'geographic_analysis': {},
            'suspicious_patterns': defaultdict(list),
        }

        # 1. Overview
        unique_ips = set(log['ip'] for log in self.logs)
        analysis['overview'] = {
            'total_requests': len(self.logs),
            'unique_ips': len(unique_ips),
            'date_range': {
                'start': min(log['datetime'] for log in self.logs),
                'end': max(log['datetime'] for log in self.logs)
            },
            'total_bandwidth': sum(log['bytes'] for log in self.logs),
            'error_rate': sum(1 for log in self.logs if log['status_code'] >= 400) / len(self.logs) * 100
        }

        # 2. Risk Scoring برای همه IP ها
        print("  📊 محاسبه امتیاز ریسک...")
        for ip in unique_ips:
            risk_info = self.calculate_ip_risk_score(ip)
            analysis['risk_scores'][ip] = risk_info
            if risk_info['risk_level'] in ['CRITICAL', 'HIGH']:
                self.suspicious_ips.add(ip)
            if risk_info['risk_level'] == 'CRITICAL':
                self.critical_ips.add(ip)

        # 3. Bot Analysis (enhanced)
        print("  🤖 تحلیل بات‌ها...")
        bot_analysis = self.analyze_bots()

        # اضافه کردن تحلیل بات‌ها به نتایج اصلی
        analysis['bot_analysis'] = {
            'legitimate': bot_analysis['legitimate'],
            'potentially_legitimate': bot_analysis['potentially_legitimate'],
            'fake': bot_analysis['fake'],
            'unknown': bot_analysis['unknown'],
            'traffic_distribution': bot_analysis['bot_traffic_distribution'],
            'activity': bot_analysis['bot_activity']
        }

        # 4. Attack Pattern Analysis
        print("  🎯 تحلیل الگوهای حمله...")
        for log in self.logs:
            request = log['request'] + ' ' + log['url']
            ip = log['ip']
            for attack_type, attack_info in self.advanced_attack_patterns.items():
                for pattern in attack_info['patterns']:
                    if re.search(pattern, request, re.IGNORECASE):
                        analysis['attack_analysis'][attack_type][ip].append({
                            'timestamp': log['timestamp'],
                            'request': log['request'][:200],
                            'status_code': log['status_code'],
                            'pattern': pattern[:50]
                        })
                        if attack_info['severity'] in ['CRITICAL', 'HIGH']:
                            self.suspicious_ips.add(ip)
                        break
                    
        # 5. Temporal Analysis
        print("  ⏰ تحلیل زمانی...")
        analysis['temporal_analysis'] = self.analyze_temporal_patterns()

        return analysis
    
    def generate_recommendations(self, analysis: Dict) -> List[str]:
        """تولید توصیه‌های امنیتی بر اساس تحلیل"""
        recommendations = []
        
        # بر اساس تعداد IP های مشکوک
        suspicious_count = len(self.suspicious_ips)
        if suspicious_count > 100:
            recommendations.append("🔴 وضعیت بحرانی: بیش از 100 IP مشکوک. فوراً فایروال را تقویت کنید.")
        elif suspicious_count > 50:
            recommendations.append("🟠 وضعیت هشدار: IP های مشکوک زیاد. بررسی و بن کردن ضروری است.")
        
        # بر اساس نوع حملات
        if 'sql_injection' in analysis['attack_analysis']:
            recommendations.append("💉 حملات SQL Injection شناسایی شد. پارامترهای ورودی را اعتبارسنجی کنید.")
        
        if 'xss' in analysis['attack_analysis']:
            recommendations.append("📝 حملات XSS شناسایی شد. خروجی‌ها را encode کنید.")
        
        if 'lfi_rfi' in analysis['attack_analysis']:
            recommendations.append("📁 حملات File Inclusion شناسایی شد. دسترسی فایل‌ها را محدود کنید.")
        
        if 'command_injection' in analysis['attack_analysis']:
            recommendations.append("💻 حملات Command Injection شناسایی شد. ورودی‌های سیستمی را فیلتر کنید.")
        
        # بر اساس بات‌های جعلی
        if analysis['bot_analysis']['fake']:
            recommendations.append(f"🤖 {len(analysis['bot_analysis']['fake'])} بات جعلی شناسایی شد. User-Agent verification فعال کنید.")
        
        # بر اساس نوع سایت
        if self.site_type == 'wordpress':
            if any('xmlrpc' in str(attack) for attack in analysis['attack_analysis'].values()):
                recommendations.append("🔒 xmlrpc.php را غیرفعال کنید.")
            recommendations.append("🛡️ پلاگین امنیتی Wordfence یا Sucuri نصب کنید.")
            recommendations.append("🔄 همه پلاگین‌ها و تم‌ها را به‌روزرسانی کنید.")
        
        elif self.site_type == 'opencart':
            recommendations.append("🔐 مسیر پنل ادمین را تغییر دهید.")
            recommendations.append("📁 دسترسی به /system/ را محدود کنید.")
            recommendations.append("🔄 همه افزونه‌ها را به‌روزرسانی کنید.")
        
        # توصیه‌های عمومی
        recommendations.append("☁️ استفاده از CDN/WAF مثل Cloudflare را در نظر بگیرید.")
        recommendations.append("📊 مانیتورینگ real-time راه‌اندازی کنید.")
        recommendations.append("🔐 2FA برای دسترسی‌های ادمین فعال کنید.")
        recommendations.append("💾 بکاپ منظم از سایت تهیه کنید.")
        
        return recommendations
    
    def export_to_excel(self, analysis: Dict, filename: str = 'security_report.xlsx'):
        """تولید گزارش Excel جامع"""
        print(f"\n📊 تولید گزارش Excel...")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            
            # 1. Overview Sheet
            overview_df = pd.DataFrame([analysis['overview']])
            overview_df.to_excel(writer, sheet_name='Overview', index=False)
            
            # 2. Risk Scores Sheet
            risk_data = []
            for ip, risk_info in analysis['risk_scores'].items():
                risk_data.append({
                    'IP': ip,
                    'Risk Score': risk_info['score'],
                    'Risk Level': risk_info['risk_level'],
                    'Request Count': risk_info['request_count'],
                    'Unique URLs': risk_info['unique_urls'],
                    'Error Count': risk_info['error_count'],
                    'Reasons': ' | '.join(risk_info['reasons'][:3])
                })
            
            risk_df = pd.DataFrame(risk_data)
            risk_df = risk_df.sort_values('Risk Score', ascending=False)
            risk_df.to_excel(writer, sheet_name='Risk Analysis', index=False)
            
            # 3. Critical IPs Sheet
            critical_data = []
            for ip in self.critical_ips:
                ip_logs = [log for log in self.logs if log['ip'] == ip]
                critical_data.append({
                    'IP': ip,
                    'Total Requests': len(ip_logs),
                    'First Seen': min(log['datetime'] for log in ip_logs),
                    'Last Seen': max(log['datetime'] for log in ip_logs),
                    'Risk Score': analysis['risk_scores'][ip]['score'],
                    'Main Reason': analysis['risk_scores'][ip]['reasons'][0] if analysis['risk_scores'][ip]['reasons'] else 'N/A'
                })
            
            if critical_data:
                critical_df = pd.DataFrame(critical_data)
                critical_df.to_excel(writer, sheet_name='Critical IPs', index=False)
            
            # 4. Attack Patterns Sheet
            attack_data = []
            for attack_type, ip_dict in analysis['attack_analysis'].items():
                for ip, attacks in ip_dict.items():
                    attack_data.append({
                        'Attack Type': attack_type.replace('_', ' ').title(),
                        'IP': ip,
                        'Count': len(attacks),
                        'Severity': self.advanced_attack_patterns[attack_type]['severity'],
                        'First Attack': attacks[0]['timestamp'] if attacks else 'N/A',
                        'Sample Request': attacks[0]['request'][:100] if attacks else 'N/A'
                    })
            
            if attack_data:
                attack_df = pd.DataFrame(attack_data)
                attack_df = attack_df.sort_values(['Severity', 'Count'], ascending=[True, False])
                attack_df.to_excel(writer, sheet_name='Attack Patterns', index=False)
            
            # 5. Bot Analysis Sheet - با شمارش درخواست بر حسب IP
            bot_data = []

            # Legitimate bots
            for company, bot_info in analysis['bot_analysis']['legitimate'].items():
                # برای هر IP جداگانه
                for ip in list(bot_info.get('ips', set()))[:100]:
                    # محاسبه تعداد درخواست این IP برای این بات
                    ip_request_count = 0
                    for log in self.logs:
                        if log['ip'] == ip:
                            ua_lower = log['user_agent'].lower()
                            if any(pattern in ua_lower for pattern in self.legitimate_bots.get(company, {}).get('patterns', [])):
                                ip_request_count += 1

                    bot_data.append({
                        'Type': 'Legitimate',
                        'Company': company,
                        'IP': ip,
                        'Verification': 'VERIFIED',
                        'Bot Type': company,
                        'Requests': ip_request_count  # تعداد واقعی درخواست این IP
                    })

            # Potentially legitimate bots
            if 'potentially_legitimate' in analysis['bot_analysis']:
                for company, bot_info in analysis['bot_analysis']['potentially_legitimate'].items():
                    for ip in list(bot_info.get('ips', set()))[:100]:
                        # محاسبه تعداد درخواست این IP
                        ip_request_count = 0
                        for log in self.logs:
                            if log['ip'] == ip:
                                ua_lower = log['user_agent'].lower()
                                if any(pattern in ua_lower for pattern in self.legitimate_bots.get(company, {}).get('patterns', [])):
                                    ip_request_count += 1

                        bot_data.append({
                            'Type': 'Potentially Legitimate',
                            'Company': company,
                            'IP': ip,
                            'Verification': 'PARTIAL',
                            'Bot Type': company,
                            'Requests': ip_request_count
                        })

            # Fake bots
            for bot_type, bot_info in analysis['bot_analysis']['fake'].items():
                for ip in list(bot_info.get('ips', set()))[:100]:
                    # محاسبه تعداد درخواست این IP
                    ip_request_count = sum(1 for log in self.logs if log['ip'] == ip)

                    bot_data.append({
                        'Type': 'FAKE',
                        'Company': bot_type,
                        'IP': ip,
                        'Verification': 'FAILED',
                        'Bot Type': 'Suspicious',
                        'Requests': ip_request_count
                    })

            # Unknown bots
            if 'unknown' in analysis['bot_analysis']:
                unknown_info = analysis['bot_analysis']['unknown']
                for ip in list(unknown_info.get('ips', set()))[:100]:
                    # محاسبه تعداد درخواست این IP
                    ip_request_count = sum(1 for log in self.logs if log['ip'] == ip)

                    bot_data.append({
                        'Type': 'Unknown',
                        'Company': 'Unknown',
                        'IP': ip,
                        'Verification': 'N/A',
                        'Bot Type': 'Unknown',
                        'Requests': ip_request_count
                    })

            if bot_data:
                bot_df = pd.DataFrame(bot_data)
                # مرتب‌سازی بر اساس Type و تعداد Requests
                bot_df = bot_df.sort_values(['Type', 'Requests'], ascending=[True, False])
                bot_df.to_excel(writer, sheet_name='Bot Analysis', index=False)
                        
            # 6. Temporal Analysis Sheet
            temporal_data = []
            for hour, count in analysis['temporal_analysis']['hourly_distribution'].items():
                temporal_data.append({
                    'Hour': f"{hour:02d}:00",
                    'Request Count': count
                })
            
            if temporal_data:
                temporal_df = pd.DataFrame(temporal_data)
                temporal_df = temporal_df.sort_values('Hour')
                temporal_df.to_excel(writer, sheet_name='Temporal Analysis', index=False)
            
            # 7. Top Requests Sheet
            request_counter = Counter()
            for log in self.logs:
                request_counter[log['url']] += 1
            
            top_requests = []
            for url, count in request_counter.most_common(100):
                top_requests.append({
                    'URL': url[:150],
                    'Count': count,
                    'Percentage': f"{(count/len(self.logs)*100):.2f}%"
                })
            
            top_requests_df = pd.DataFrame(top_requests)
            top_requests_df.to_excel(writer, sheet_name='Top Requests', index=False)
            
            # 8. Bot Error Analysis Sheet
            bot_error_data = []
            for log in self.logs:
                if log['status_code'] >= 400:
                    # بررسی اینکه آیا این User-Agent بات است
                    ua_lower = log['user_agent'].lower()
                    is_bot = False
                    bot_type = None 

                    # بررسی بات‌های معتبر
                    for company, bot_info in self.legitimate_bots.items():
                        if any(pattern in ua_lower for pattern in bot_info['patterns']):
                            is_bot = True
                            bot_type = f"Legitimate {company}"
                            break   

                    # بررسی بات‌های جعلی
                    if not is_bot:
                        if any(tool in ua_lower for tool in self.suspicious_user_agents['hacking_tools']):
                            is_bot = True
                            bot_type = "Hacking Tool"
                        elif any(bot in ua_lower for bot in ['bot', 'crawler', 'spider']):
                            is_bot = True
                            bot_type = "Suspicious Bot" 

                    if is_bot:
                        bot_error_data.append({
                            'IP': log['ip'],
                            'Status Code': log['status_code'],
                            'Error Type': self.error_codes.get(str(log['status_code']), 'Unknown'),
                            'URL': log['url'][:100],
                            'Timestamp': log['timestamp'],
                            'User Agent': log['user_agent'][:150],
                            'Bot Type': bot_type
                        })  

            if bot_error_data[:1000]:  # محدود به 1000 رکورد
                bot_error_df = pd.DataFrame(bot_error_data[:1000])
                bot_error_df.to_excel(writer, sheet_name='Bot Error Analysis', index=False)
            
            # 9. User Agent Analysis Sheet
            ua_counter = Counter(log['user_agent'] for log in self.logs)
            ua_data = []
            
            for ua, count in ua_counter.most_common(50):
                ua_type = 'Unknown'
                if ua == '-' or len(ua) < 5:
                    ua_type = 'Empty/Invalid'
                elif any(tool in ua.lower() for tool in self.suspicious_user_agents['hacking_tools']):
                    ua_type = 'Hacking Tool'
                elif any(bot in ua.lower() for bot in ['bot', 'crawler', 'spider']):
                    ua_type = 'Bot/Crawler'
                else:
                    try:
                        parsed = parse(ua)
                        if parsed.is_mobile:
                            ua_type = 'Mobile'
                        elif parsed.is_tablet:
                            ua_type = 'Tablet'
                        elif parsed.is_pc:
                            ua_type = 'Desktop'
                    except:
                        pass
                    
                ua_data.append({
                    'User Agent': ua[:150],
                    'Count': count,
                    'Type': ua_type,
                    'Percentage': f"{(count/len(self.logs)*100):.2f}%"
                })
            
            ua_df = pd.DataFrame(ua_data)
            ua_df.to_excel(writer, sheet_name='User Agents', index=False)
            
            # 10. Bot Statistics Sheet - اضافه شده
            bot_stats_data = []
            
            # آمار بات‌های معتبر
            for company, bot_info in analysis['bot_analysis']['legitimate'].items():
                if bot_info.get('requests', 0) > 0:
                    bot_stats_data.append({
                        'Bot Name': company,
                        'Type': 'Legitimate',
                        'Total Requests': bot_info.get('requests', 0),
                        'Unique IPs': bot_info.get('ips_count', len(bot_info.get('ips', []))),
                        'Unique URLs': bot_info.get('unique_urls_count', len(bot_info.get('unique_urls', []))),
                        'First Seen': bot_info.get('first_seen', 'N/A'),
                        'Last Seen': bot_info.get('last_seen', 'N/A')
                    })
            
            # آمار بات‌های احتمالی
            if 'potentially_legitimate' in analysis['bot_analysis']:
                for company, bot_info in analysis['bot_analysis']['potentially_legitimate'].items():
                    if bot_info.get('requests', 0) > 0:
                        bot_stats_data.append({
                            'Bot Name': company,
                            'Type': 'Potentially Legitimate',
                            'Total Requests': bot_info.get('requests', 0),
                            'Unique IPs': bot_info.get('ips_count', len(bot_info.get('ips', []))),
                            'Unique URLs': bot_info.get('unique_urls_count', len(bot_info.get('unique_urls', []))),
                            'First Seen': bot_info.get('first_seen', 'N/A'),
                            'Last Seen': bot_info.get('last_seen', 'N/A')
                        })
            
            # آمار بات‌های جعلی
            for bot_type, bot_info in analysis['bot_analysis']['fake'].items():
                if bot_info.get('requests', 0) > 0:
                    bot_stats_data.append({
                        'Bot Name': bot_type,
                        'Type': 'Fake/Suspicious',
                        'Total Requests': bot_info.get('requests', 0),
                        'Unique IPs': bot_info.get('ips_count', len(bot_info.get('ips', []))),
                        'Unique URLs': 'N/A',
                        'First Seen': bot_info.get('first_seen', 'N/A'),
                        'Last Seen': bot_info.get('last_seen', 'N/A')
                    })
            
            if bot_stats_data:
                bot_stats_df = pd.DataFrame(bot_stats_data)
                bot_stats_df = bot_stats_df.sort_values(['Type', 'Total Requests'], ascending=[True, False])
                bot_stats_df.to_excel(writer, sheet_name='Bot Statistics', index=False)
            
            # Formatting
            workbook = writer.book
            
            # اضافه کردن استایل به همه sheet ها
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Conditional formatting for Risk Analysis sheet
                if sheet_name == 'Risk Analysis':
                    for row in range(2, len(risk_df) + 2):
                        risk_level = worksheet[f'C{row}'].value
                        if risk_level == 'CRITICAL':
                            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        elif risk_level == 'HIGH':
                            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                        elif risk_level == 'MEDIUM':
                            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        elif risk_level == 'LOW':
                            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        else:
                            fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        
                        worksheet[f'C{row}'].fill = fill
                
                # Conditional formatting for Bot Analysis sheet
                if sheet_name == 'Bot Analysis':
                    for row in range(2, worksheet.max_row + 1):
                        bot_type = worksheet[f'A{row}'].value
                        if bot_type == 'FAKE':
                            fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                            for col in range(1, 7):  # تمام ستون‌ها
                                worksheet.cell(row=row, column=col).fill = fill
                        elif bot_type == 'Potentially Legitimate':
                            fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
                            for col in range(1, 7):
                                worksheet.cell(row=row, column=col).fill = fill
                        elif bot_type == 'Legitimate':
                            fill = PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")
                            for col in range(1, 7):
                                worksheet.cell(row=row, column=col).fill = fill
        
        print(f"✅ گزارش Excel در {filename} ذخیره شد")
        return filename
    
    def export_firewall_rules(self):
        """تولید قوانین فایروال برای سیستم‌های مختلف"""
        
        # 2. .htaccess for Apache
        with open('htaccess_rules.txt', 'w') as f:
            f.write(f"# Security Rules - Generated: {datetime.now()}\n")
            f.write(f"# Add to your .htaccess file\n\n")
            
            f.write("# Block suspicious IPs\n")
            f.write("order allow,deny\n")
            
            for ip in sorted(self.suspicious_ips):
                f.write(f"deny from {ip}\n")
            
            f.write("allow from all\n\n")
            
            # Additional security headers
            f.write("# Security Headers\n")
            f.write("Header set X-Frame-Options \"SAMEORIGIN\"\n")
            f.write("Header set X-Content-Type-Options \"nosniff\"\n")
            f.write("Header set X-XSS-Protection \"1; mode=block\"\n")
        
        # 3. nginx configuration
        with open('nginx_rules.conf', 'w') as f:
            f.write(f"# Nginx Security Rules - Generated: {datetime.now()}\n\n")
            
            f.write("# Block suspicious IPs\n")
            for ip in sorted(self.suspicious_ips):
                f.write(f"deny {ip};\n")
            
            f.write("\n# Rate limiting\n")
            f.write("limit_req_zone $binary_remote_addr zone=one:10m rate=10r/s;\n")
            f.write("limit_conn_zone $binary_remote_addr zone=addr:10m;\n")
        
        # 4. CSF (ConfigServer Firewall)
        with open('csf_deny.txt', 'w') as f:
            f.write(f"# CSF Deny List - Generated: {datetime.now()}\n")
            f.write("# Copy to: /etc/csf/csf.deny\n\n")
            
            for ip in sorted(self.suspicious_ips):
                risk_score = self.analysis_results.get('risk_scores', {}).get(ip, {}).get('score', 0)
                f.write(f"{ip} # Risk Score: {risk_score}\n")
        
        # 5. fail2ban configuration
        with open('fail2ban_jail.conf', 'w') as f:
            f.write(f"# Fail2ban Jail Configuration - Generated: {datetime.now()}\n\n")
            
            f.write("[web-security]\n")
            f.write("enabled = true\n")
            f.write("filter = web-security\n")
            f.write("action = iptables-multiport[name=WebSecurity, port=\"http,https\"]\n")
            f.write("logpath = /var/log/apache2/access.log\n")
            f.write("maxretry = 3\n")
            f.write("findtime = 600\n")
            f.write("bantime = 86400\n")
        
        print("✅ قوانین فایروال تولید شد:")
        print("  • iptables_rules.sh")
        print("  • htaccess_rules.txt")
        print("  • nginx_rules.conf")
        print("  • csf_deny.txt")
        print("  • fail2ban_jail.conf")
    
    def generate_report(self):
        """تولید گزارش نهایی"""
        print("" + "="*80)
        print(f"📊 گزارش تحلیل امنیتی پیشرفته")
        print(f"🕐 زمان: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"📌 نوع سایت: {self.site_type.upper()}")
        print("="*80)

        # اجرای تحلیل جامع
        self.analysis_results = self.comprehensive_analysis()

        # 1. خلاصه وضعیت
        print("#### 📋 خلاصه وضعیت امنیتی")
        print("-" * 60)
        overview = self.analysis_results['overview']
        print(f"  • کل درخواست‌ها: {overview['total_requests']:,}")
        print(f"  • IP های یکتا: {overview['unique_ips']:,}")
        print(f"  • IP های مشکوک: {len(self.suspicious_ips)} ({len(self.suspicious_ips)/overview['unique_ips']*100:.1f}%)")
        print(f"  • IP های بحرانی: {len(self.critical_ips)}")
        print(f"  • حجم ترافیک: {overview['total_bandwidth']/1024/1024:.2f} MB")
        print(f"  • نرخ خطا: {overview['error_rate']:.2f}%")

        # 2. تهدیدات بحرانی
        critical_threats = [ip for ip, info in self.analysis_results['risk_scores'].items() 
                          if info['risk_level'] == 'CRITICAL']
        if critical_threats:
            print("#### 🚨 تهدیدات بحرانی (نیاز به اقدام فوری)")
            print("-" * 60)
            for ip in critical_threats[:10]:
                risk_info = self.analysis_results['risk_scores'][ip]
                print(f"  🔴 {ip}")
                print(f"     امتیاز ریسک: {risk_info['score']}")
                print(f"     دلایل: {', '.join(risk_info['reasons'][:3])}")

        # 3. تحلیل بات‌ها و زمان بازدید آنها
        print("#### 🤖 تحلیل بات‌ها و زمان بازدید")
        print("-" * 60)

        # بات‌های معتبر
        legitimate_bots = self.analysis_results['bot_analysis']['legitimate']
        if legitimate_bots:
            print("✅ بات‌های معتبر:")
            for bot_type, data in legitimate_bots.items():
                if data['requests'] > 0:
                    print(f"**{bot_type}:**")
                    print(f"      • تعداد درخواست‌ها: {data['requests']:,}")
                    print(f"      • IP های منحصربه‌فرد: {data['ips_count']}")
                    print(f"      • اولین بازدید: {data['first_seen'].strftime('%Y-%m-%d %H:%M:%S') if data['first_seen'] else 'N/A'}")
                    print(f"      • آخرین بازدید: {data['last_seen'].strftime('%Y-%m-%d %H:%M:%S') if data['last_seen'] else 'N/A'}")
                    print(f"      • مسیرهای پربازدید: {', '.join([url for url, _ in data['top_urls'][:3]])}")

        # بات‌های احتمالاً معتبر
        potentially_legit = self.analysis_results['bot_analysis']['potentially_legitimate']
        if potentially_legit:
            print("🟡 بات‌های احتمالاً معتبر (نیاز به بررسی):")
            for bot_type, data in potentially_legit.items():
                if data['requests'] > 0:
                    print(f"**{bot_type}:**")
                    print(f"      • تعداد درخواست‌ها: {data['requests']:,}")
                    print(f"      • IP های منحصربه‌فرد: {data['ips_count']}")
                    print(f"      • دلیل: User-Agent معتبر اما IP range تأیید نشده")

        # بات‌های جعلی
        fake_bots = self.analysis_results['bot_analysis']['fake']
        if fake_bots:
            print("❌ بات‌های جعلی شناسایی شده:")
            for bot_type, data in fake_bots.items():
                if data['requests'] > 0:
                    print(f"**{bot_type}:**")
                    print(f"      • تعداد درخواست‌ها: {data['requests']:,}")
                    print(f"      • IP های منحصربه‌فرد: {data['ips_count']}")
                    print(f"      • User-Agentهای متداول: {', '.join([ua for ua, _ in data['top_user_agents'][:3]])}")
                    print(f"      • مسیرهای پربازدید: {', '.join([url for url, _ in data['top_urls'][:3]])}")

        # بات‌های ناشناس
        unknown_bots = self.analysis_results['bot_analysis']['unknown']
        if unknown_bots['requests'] > 0:
            print("⚪ بات‌های ناشناس:")
            print(f"  • تعداد درخواست‌ها: {unknown_bots['requests']:,}")
            print(f"  • IP های منحصربه‌فرد: {unknown_bots['ips_count']}")
            print(f"  • User-Agentهای متداول: {', '.join([ua for ua, _ in unknown_bots['top_user_agents'][:3]])}")
            print(f"  • مسیرهای پربازدید: {', '.join([url for url, _ in unknown_bots['top_urls'][:3]])}")

        # 4. حملات شناسایی شده
        if self.analysis_results['attack_analysis']:
            print("#### 🎯 حملات شناسایی شده")
            print("-" * 60)
            attack_summary = Counter()
            for attack_type, ip_dict in self.analysis_results['attack_analysis'].items():
                attack_summary[attack_type] = len(ip_dict)
            for attack_type, count in attack_summary.most_common():
                severity = self.advanced_attack_patterns[attack_type]['severity']
                icon = {'CRITICAL': '🔴', 'HIGH': '🟠', 'MEDIUM': '🟡', 'LOW': '🟢'}.get(severity, '⚪')
                print(f"  {icon} {attack_type.replace('_', ' ').title()}: {count} IP")

        # 5. الگوهای زمانی
        temporal = self.analysis_results['temporal_analysis']
        if temporal.get('peak_hours'):
            print("#### ⏰ الگوهای زمانی")
            print("-" * 60)
            print("  ساعات پرترافیک:")
            for hour, count in temporal['peak_hours'][:5]:
                print(f"    • {hour:02d}:00 - {count:,} درخواست")


        # 7. آمار نهایی
        print("#### 📈 آمار نهایی")
        print("-" * 60)
        # توزیع سطح ریسک
        risk_distribution = Counter(info['risk_level'] for info in self.analysis_results['risk_scores'].values())
        print("  توزیع سطح ریسک IP ها:")
        for level in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'SAFE']:
            count = risk_distribution.get(level, 0)
            if count > 0:
                percentage = count / overview['unique_ips'] * 100
                bar = '█' * int(percentage / 2)
                print(f"    {level:8s}: {count:4d} ({percentage:5.1f}%) {bar}")

        return self.analysis_results
    
    def analyze_bot_visit_times(self) -> Dict:
        """تحلیل زمان بازدید بات‌های مختلف"""
        bot_visits = {
            'legitimate_bot_times': defaultdict(lambda: {
                'visits': [],
                'total_visits': 0,
                'first_visit': None,
                'last_visit': None,
                'hourly_distribution': defaultdict(int),
                'daily_distribution': defaultdict(int),
                'weekly_distribution': defaultdict(int),
                'peak_hours': [],
                'off_hours_visits': 0,  # بازدید در ساعات غیرعادی
                'unique_ips': set(),
                'unique_urls': set(),
                'average_requests_per_ip': 0,
                'crawl_rate': 0,  # تعداد درخواست در دقیقه
                'response_codes': defaultdict(int)
            }),
            'fake_bot_times': defaultdict(lambda: {
                'visits': [],
                'claimed': '',
                'reason': '',
                'user_agents': set(),
                'ips': set()
            }),
            'potentially_legitimate': defaultdict(lambda: {
                'visits': [],
                'total_visits': 0,
                'unique_ips': set(),
                'verification_status': 'partial'
            }),
            'bot_behavior_analysis': {
                'crawl_patterns': defaultdict(list),
                'suspicious_patterns': [],
                'bot_comparison': {}
            }
        }
        
        # بررسی وجود dnspython
        DNSPYTHON_AVAILABLE = False
        dns_resolver = None
        try:
            import dns.resolver
            import dns.reversename
            DNSPYTHON_AVAILABLE = True
            dns_resolver = dns.resolver.Resolver()
            dns_resolver.timeout = 0.3
            dns_resolver.lifetime = 1.0
            dns_resolver.nameservers = ['8.8.8.8', '8.8.4.4']
        except ImportError:
            DNSPYTHON_AVAILABLE = False
        
        # کش DNS برای بهبود performance
        dns_cache = {}
        
        # پردازش همه لاگ‌ها
        for log in self.logs:
            ip = log['ip']
            ua = log['user_agent']
            url = log['url']
            dt = log['datetime']
            status_code = log['status_code']
            ua_lower = ua.lower()
            
            # بررسی بات گوگل با GoogleBotVerifier
            google_result = self.google_verifier.verify_google_bot(ip, ua)
            if google_result['is_google']:
                bot_type = 'Google'
                bot_visits['legitimate_bot_times'][bot_type]['visits'].append(dt)
                bot_visits['legitimate_bot_times'][bot_type]['hourly_distribution'][dt.hour] += 1
                bot_visits['legitimate_bot_times'][bot_type]['daily_distribution'][dt.weekday()] += 1
                bot_visits['legitimate_bot_times'][bot_type]['unique_ips'].add(ip)
                bot_visits['legitimate_bot_times'][bot_type]['unique_urls'].add(url)
                bot_visits['legitimate_bot_times'][bot_type]['response_codes'][status_code] += 1
                
                # بررسی بازدید در ساعات غیرعادی (بین 2 تا 6 صبح)
                if 2 <= dt.hour <= 6:
                    bot_visits['legitimate_bot_times'][bot_type]['off_hours_visits'] += 1
                continue
            
            # بررسی سایر بات‌های معتبر
            bot_identified = False
            for company, bot_info in self.legitimate_bots.items():
                if company == 'Google':  # قبلاً بررسی شده
                    continue
                    
                if not bot_info['patterns']:
                    continue
                
                # بررسی User-Agent
                ua_matches = any(pattern in ua_lower for pattern in bot_info['patterns'])
                
                if not ua_matches:
                    continue
                
                # بررسی IP range
                ip_matches = False
                if bot_info['ip_ranges']:
                    try:
                        ip_obj = ipaddress.ip_address(ip)
                        for ip_range in bot_info['ip_ranges']:
                            if '/' in ip_range:
                                network = ipaddress.ip_network(ip_range, strict=False)
                                if ip_obj in network:
                                    ip_matches = True
                                    break
                            else:
                                if ip.startswith(ip_range.split('/')[0]):
                                    ip_matches = True
                                    break
                    except Exception:
                        pass
                    
                # اگر IP range وجود دارد و مطابقت دارد
                if bot_info['ip_ranges'] and ip_matches:
                    # بات کاملاً معتبر
                    bot_visits['legitimate_bot_times'][company]['visits'].append(dt)
                    bot_visits['legitimate_bot_times'][company]['hourly_distribution'][dt.hour] += 1
                    bot_visits['legitimate_bot_times'][company]['daily_distribution'][dt.weekday()] += 1
                    bot_visits['legitimate_bot_times'][company]['unique_ips'].add(ip)
                    bot_visits['legitimate_bot_times'][company]['unique_urls'].add(url)
                    bot_visits['legitimate_bot_times'][company]['response_codes'][status_code] += 1
                    
                    if 2 <= dt.hour <= 6:
                        bot_visits['legitimate_bot_times'][company]['off_hours_visits'] += 1
                    
                    bot_identified = True
                    break
                    
                # اگر IP range وجود ندارد، با DNS بررسی کنیم
                elif not bot_info['ip_ranges'] and bot_info['dns_suffix']:
                    # DNS verification
                    hostname = None
                    if ip in dns_cache:
                        hostname = dns_cache[ip]
                    #else:
                    #    if DNSPYTHON_AVAILABLE and dns_resolver:
                    #        try:
                    #            rev_name = dns.reversename.from_address(ip)
                    #            hostname = str(dns_resolver.resolve(rev_name, "PTR")[0])
                    #            dns_cache[ip] = hostname
                    #        except:
                    #            dns_cache[ip] = None
                    #    else:
                    #        try:
                    #            hostname = socket.gethostbyaddr(ip)[0]
                    #            dns_cache[ip] = hostname
                    #        except:
                    #            dns_cache[ip] = None
                    
                    if hostname:
                        dns_matches = any(hostname.endswith(suffix) for suffix in bot_info['dns_suffix'])
                        if dns_matches:
                            # بات معتبر با DNS verification
                            bot_visits['legitimate_bot_times'][company]['visits'].append(dt)
                            bot_visits['legitimate_bot_times'][company]['hourly_distribution'][dt.hour] += 1
                            bot_visits['legitimate_bot_times'][company]['daily_distribution'][dt.weekday()] += 1
                            bot_visits['legitimate_bot_times'][company]['unique_ips'].add(ip)
                            bot_visits['legitimate_bot_times'][company]['unique_urls'].add(url)
                            bot_visits['legitimate_bot_times'][company]['response_codes'][status_code] += 1
                            
                            if 2 <= dt.hour <= 6:
                                bot_visits['legitimate_bot_times'][company]['off_hours_visits'] += 1
                            
                            bot_identified = True
                            break
                        else:
                            # User-Agent درست اما DNS تأیید نشد - مشکوک
                            bot_visits['fake_bot_times'][ip]['visits'].append(dt)
                            bot_visits['fake_bot_times'][ip]['claimed'] = company
                            bot_visits['fake_bot_times'][ip]['reason'] = 'DNS verification failed'
                            bot_visits['fake_bot_times'][ip]['user_agents'].add(ua)
                            bot_visits['fake_bot_times'][ip]['ips'].add(ip)
                            bot_identified = True
                            break
                        
                # اگر IP range وجود دارد اما مطابقت ندارد
                elif bot_info['ip_ranges'] and not ip_matches:
                    # احتمالاً بات جعلی یا potentially legitimate
                    bot_visits['potentially_legitimate'][company]['visits'].append(dt)
                    bot_visits['potentially_legitimate'][company]['total_visits'] += 1
                    bot_visits['potentially_legitimate'][company]['unique_ips'].add(ip)
                    bot_visits['potentially_legitimate'][company]['verification_status'] = 'UA matches, IP doesn\'t match'
                    bot_identified = True
                    break
                
            # بررسی بات‌های مشکوک و ابزارهای هک
            if not bot_identified:
                # بررسی ابزارهای هک
                for tool in self.suspicious_user_agents['hacking_tools']:
                    if tool in ua_lower:
                        bot_visits['fake_bot_times'][ip]['visits'].append(dt)
                        bot_visits['fake_bot_times'][ip]['claimed'] = 'Hacking Tool'
                        bot_visits['fake_bot_times'][ip]['reason'] = f'Contains: {tool}'
                        bot_visits['fake_bot_times'][ip]['user_agents'].add(ua)
                        bot_visits['fake_bot_times'][ip]['ips'].add(ip)
                        bot_identified = True
                        break
                    
                # بررسی User-Agent خالی یا مشکوک
                if not bot_identified and (ua == '-' or len(ua) < 5):
                    bot_visits['fake_bot_times'][ip]['visits'].append(dt)
                    bot_visits['fake_bot_times'][ip]['claimed'] = 'Invalid UA'
                    bot_visits['fake_bot_times'][ip]['reason'] = 'Empty or too short User-Agent'
                    bot_visits['fake_bot_times'][ip]['user_agents'].add(ua)
                    bot_visits['fake_bot_times'][ip]['ips'].add(ip)
        
        # محاسبه آمار نهایی برای بات‌های معتبر
        for company, data in bot_visits['legitimate_bot_times'].items():
            if data['visits']:
                data['total_visits'] = len(data['visits'])
                data['first_visit'] = min(data['visits'])
                data['last_visit'] = max(data['visits'])
                
                # محاسبه ساعات پربازدید
                peak_hours = sorted(data['hourly_distribution'].items(), 
                                  key=lambda x: x[1], reverse=True)
                data['peak_hours'] = [hour for hour, _ in peak_hours[:5]]
                
                # محاسبه میانگین درخواست برای هر IP
                if data['unique_ips']:
                    data['average_requests_per_ip'] = data['total_visits'] / len(data['unique_ips'])
                
                # محاسبه نرخ crawl (درخواست در دقیقه)
                if data['first_visit'] and data['last_visit']:
                    time_diff = (data['last_visit'] - data['first_visit']).total_seconds() / 60
                    if time_diff > 0:
                        data['crawl_rate'] = data['total_visits'] / time_diff
        
        # تحلیل رفتار بات‌ها
        bot_visits['bot_behavior_analysis'] = self._analyze_bot_behavior(bot_visits)
        
        return bot_visits
    
    def _analyze_bot_behavior(self, bot_visits: Dict) -> Dict:
        """تحلیل رفتار و الگوی بازدید بات‌ها"""
        behavior_analysis = {
            'crawl_patterns': {},
            'suspicious_patterns': [],
            'bot_comparison': {},
        }
        
        # تحلیل الگوی crawl برای هر بات معتبر
        for company, data in bot_visits['legitimate_bot_times'].items():
            if data['total_visits'] > 0:
                pattern = {
                    'regularity': 'regular' if data['crawl_rate'] < 10 else 'aggressive',
                    'peak_activity': f"{data['peak_hours'][0]:02d}:00" if data['peak_hours'] else 'N/A',
                    'off_hours_percentage': (data['off_hours_visits'] / data['total_visits']) * 100,
                    'error_rate': (data['response_codes'].get(404, 0) / data['total_visits']) * 100,
                    'urls_per_ip': len(data['unique_urls']) / len(data['unique_ips']) if data['unique_ips'] else 0
                }
                behavior_analysis['crawl_patterns'][company] = pattern
                
                # شناسایی الگوهای مشکوک
                if pattern['regularity'] == 'aggressive':
                    behavior_analysis['suspicious_patterns'].append(
                        f"{company}: Aggressive crawling rate ({data['crawl_rate']:.1f} req/min)"
                    )
                
                if pattern['error_rate'] > 20:
                    behavior_analysis['suspicious_patterns'].append(
                        f"{company}: High error rate ({pattern['error_rate']:.1f}%)"
                    )
        
        # مقایسه بات‌ها
        if bot_visits['legitimate_bot_times']:
            total_bot_visits = sum(data['total_visits'] 
                                  for data in bot_visits['legitimate_bot_times'].values())
            for company, data in bot_visits['legitimate_bot_times'].items():
                if data['total_visits'] > 0:
                    behavior_analysis['bot_comparison'][company] = {
                        'percentage_of_bot_traffic': (data['total_visits'] / total_bot_visits) * 100,
                        'unique_ips': len(data['unique_ips']),
                        'unique_urls': len(data['unique_urls'])
                    }
        
        return behavior_analysis
    
    def export_json_report(self, filename: str = 'security_report.json'):
        """ذخیره گزارش به فرمت JSON با رفع مشکل encoding"""
        report = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'site_type': self.site_type,
                'log_file': self.log_file_path,
                'analysis_version': '2.0'
            },
            'statistics': {
                'total_requests': len(self.logs),
                'unique_ips': len(set(log['ip'] for log in self.logs)),
                'suspicious_ips': len(self.suspicious_ips),
                'critical_ips': len(self.critical_ips),
                'fake_bots': len(self.analysis_results['bot_analysis']['fake'])
            },
            'threats': {
                'critical_ips': list(self.critical_ips),
                'suspicious_ips': list(self.suspicious_ips),
                'fake_bots': dict(self.analysis_results['bot_analysis']['fake'])
            },
            'bot_visit_times': self.analyze_bot_visit_times(),
            'risk_scores': {
                ip: {
                    'score': info['score'],
                    'level': info['risk_level'],
                    'reasons': info['reasons']
                }
                for ip, info in sorted(
                    self.analysis_results['risk_scores'].items(),
                    key=lambda x: x[1]['score'],
                    reverse=True
                )[:100]  # Top 100 risky IPs
            },
            'attack_patterns': {
                attack_type: {
                    'severity': self.advanced_attack_patterns[attack_type]['severity'],
                    'affected_ips': list(ip_dict.keys())[:20]
                }
                for attack_type, ip_dict in self.analysis_results['attack_analysis'].items()
            },
        }
        
        # ذخیره با encoding UTF-8
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n✅ گزارش JSON در {filename} ذخیره شد")
        return filename

    def generate_bot_timeline_report(self) -> Dict:
        """تولید گزارش خط زمانی (Timeline) برای بازدید بات‌ها"""
        from collections import defaultdict
        from datetime import datetime, timedelta
        import json

        print("\n⏰ تولید گزارش خط زمانی بات‌ها...")

        timeline = {
            'search_engines': defaultdict(list),  # Google, Bing, etc.
            'ai_bots': defaultdict(list),  # OpenAI, Perplexity, etc.
            'social_media': defaultdict(list),  # Facebook, LinkedIn, etc.
            'seo_tools': defaultdict(list),  # Ahrefs, SemRush, etc.
            'other_bots': defaultdict(list),
            'timeline_events': [],  # همه رویدادها به صورت chronological
            'hourly_summary': defaultdict(lambda: defaultdict(int)),
            'daily_summary': defaultdict(lambda: defaultdict(int)),
            'bot_page_analysis': defaultdict(lambda: defaultdict(list))
        }

        # دسته‌بندی بات‌ها
        bot_categories = {
            'search_engines': ['Google', 'Bing', 'Yandex', 'Baidu', 'DuckDuckGo'],
            'ai_bots': ['OpenAI', 'PerplexityBot', 'PerplexityUser', 'Cohere', 'Mistral', 
                        'AllenInstitute', 'YouCom'],
            'social_media': ['Meta', 'LinkedIn', 'ByteDance'],
            'seo_tools': ['SemRush', 'Ahrefs'],
            'other_bots': ['Amazon', 'Apple', 'CommonCrawl', 'Diffbot', 'Omgili', 'Timpi']
        }

        # کش برای بررسی سریع‌تر
        processed_entries = set()

        # پردازش لاگ‌ها
        for log in sorted(self.logs, key=lambda x: x['datetime']):
            ip = log['ip']
            ua = log['user_agent']
            url = log['url']
            dt = log['datetime']
            status_code = log['status_code']
            ua_lower = ua.lower()

            # کلید یکتا برای جلوگیری از تکرار
            entry_key = f"{dt}_{ip}_{url}"
            if entry_key in processed_entries:
                continue
            processed_entries.add(entry_key)

            # تشخیص نوع بات
            bot_identified = False
            bot_name = None
            bot_category = None
            bot_subtype = None

            # بررسی بات گوگل با جزئیات بیشتر
            google_result = self.google_verifier.verify_google_bot(ip, ua)
            if google_result['is_google']:
                bot_name = 'Google'
                bot_category = 'search_engines'

                # تشخیص نوع دقیق بات گوگل
                if 'googlebot' in ua_lower:
                    if 'googlebot-image' in ua_lower:
                        bot_subtype = 'Googlebot-Image'
                    elif 'googlebot-video' in ua_lower:
                        bot_subtype = 'Googlebot-Video'
                    elif 'googlebot-news' in ua_lower:
                        bot_subtype = 'Googlebot-News'
                    elif 'smartphone' in ua_lower:
                        bot_subtype = 'Googlebot-Mobile'
                    else:
                        bot_subtype = 'Googlebot'
                elif 'adsbot-google' in ua_lower:
                    bot_subtype = 'AdsBot-Google'
                elif 'mediapartners-google' in ua_lower:
                    bot_subtype = 'Mediapartners-Google'
                elif 'google-inspectiontool' in ua_lower:
                    bot_subtype = 'Google-InspectionTool'
                elif 'google-extended' in ua_lower:
                    bot_subtype = 'Google-Extended'
                elif 'google-cloudvertexbot' in ua_lower:
                    bot_subtype = 'Google-CloudVertexBot'
                elif 'googleother' in ua_lower:
                    bot_subtype = 'GoogleOther'
                else:
                    bot_subtype = 'Google-Bot'

                bot_identified = True

            # بررسی سایر بات‌ها
            if not bot_identified:
                for company, bot_info in self.legitimate_bots.items():
                    if company == 'Google':
                        continue
                    
                    if not bot_info['patterns']:
                        continue
                    
                    if any(pattern in ua_lower for pattern in bot_info['patterns']):
                        bot_name = company

                        # تعیین دسته
                        for cat, bot_list in bot_categories.items():
                            if company in bot_list:
                                bot_category = cat
                                break
                            
                        if not bot_category:
                            bot_category = 'other_bots'

                        # تعیین زیرنوع برای بات‌های خاص
                        if company == 'OpenAI':
                            if 'gptbot' in ua_lower:
                                bot_subtype = 'GPTBot'
                            elif 'chatgpt-user' in ua_lower:
                                bot_subtype = 'ChatGPT-User'
                            elif 'oai-searchbot' in ua_lower:
                                bot_subtype = 'OAI-SearchBot'
                            else:
                                bot_subtype = 'OpenAI-Bot'
                        elif company == 'PerplexityBot':
                            bot_subtype = 'PerplexityBot'
                        elif company == 'PerplexityUser':
                            bot_subtype = 'Perplexity-User'
                        else:
                            bot_subtype = company

                        bot_identified = True
                        break
                    
            if bot_identified and bot_name and bot_category:
                # ایجاد رویداد timeline
                event = {
                    'timestamp': dt.isoformat(),
                    'datetime_obj': dt,
                    'bot_name': bot_name,
                    'bot_type': bot_subtype or bot_name,
                    'category': bot_category,
                    'ip': ip,
                    'url': url,
                    'status_code': status_code,
                    'user_agent': ua[:100]  # محدود کردن طول
                }

                # اضافه کردن به timeline اصلی
                timeline['timeline_events'].append(event)

                # اضافه کردن به دسته مربوطه
                timeline[bot_category][bot_name].append(event)

                # آمار ساعتی و روزانه
                hour_key = dt.strftime('%Y-%m-%d %H:00')
                day_key = dt.strftime('%Y-%m-%d')

                timeline['hourly_summary'][hour_key][bot_subtype or bot_name] += 1
                timeline['daily_summary'][day_key][bot_subtype or bot_name] += 1

                # تحلیل صفحات بازدید شده توسط هر بات
                timeline['bot_page_analysis'][bot_subtype or bot_name][url].append({
                    'timestamp': dt.isoformat(),
                    'ip': ip,
                    'status': status_code
                })

        # تولید گزارش HTML زیبا
        self._generate_timeline_html(timeline)

        # تولید گزارش Text زیبا
        self._generate_timeline_text(timeline)

        # تولید گزارش JSON
        self._generate_timeline_json(timeline)

        return timeline 

    def _generate_timeline_html(self, timeline: Dict):
        """تولید گزارش HTML از Timeline"""
        
        html_content = """
        <!DOCTYPE html>
    <html dir="rtl" lang="fa">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Bot Timeline Analysis Report</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: #333;
                padding: 20px;
                direction: ltr;
            }
            .container {
                max-width: 1400px;
                margin: 0 auto;
                background: white;
                border-radius: 15px;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                padding: 30px;
            }
            h1 {
                text-align: center;
                color: #764ba2;
                margin-bottom: 30px;
                font-size: 2.5em;
            }
            h2 {
                color: #667eea;
                margin: 30px 0 20px;
                border-bottom: 3px solid #667eea;
                padding-bottom: 10px;
            }
            h3 {
                color: #555;
                margin: 20px 0 10px;
            }
            .summary-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 20px;
                margin: 30px 0;
            }
            .summary-card {
                background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
                padding: 20px;
                border-radius: 10px;
                box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            }
            .summary-card h3 {
                color: #667eea;
                margin-bottom: 15px;
            }
            .timeline {
                position: relative;
                padding: 20px 0;
            }
            .timeline-item {
                display: flex;
                margin-bottom: 20px;
                position: relative;
                padding-left: 40px;
            }
            .timeline-item::before {
                content: '';
                position: absolute;
                left: 10px;
                top: 5px;
                width: 12px;
                height: 12px;
                border-radius: 50%;
                background: #667eea;
            }
            .timeline-item::after {
                content: '';
                position: absolute;
                left: 15px;
                top: 17px;
                bottom: -20px;
                width: 2px;
                background: #ddd;
            }
            .timeline-item:last-child::after {
                display: none;
            }
            .timeline-time {
                min-width: 150px;
                color: #888;
                font-weight: bold;
            }
            .timeline-content {
                flex: 1;
                background: #f8f9fa;
                padding: 15px;
                border-radius: 8px;
                margin-left: 20px;
                box-shadow: 0 3px 10px rgba(0,0,0,0.1);
            }
            .bot-google { background: linear-gradient(135deg, #4285F4, #34A853); }
            .bot-bing { background: linear-gradient(135deg, #00BCF2, #0078D4); }
            .bot-openai { background: linear-gradient(135deg, #10A37F, #1A7F64); }
            .bot-perplexity { background: linear-gradient(135deg, #8B5CF6, #7C3AED); }
            .bot-meta { background: linear-gradient(135deg, #1877F2, #42B3F4); }
            .bot-badge {
                display: inline-block;
                padding: 3px 10px;
                border-radius: 15px;
                color: white;
                font-size: 0.85em;
                font-weight: bold;
                margin-right: 10px;
            }
            .url-path {
                color: #666;
                font-family: monospace;
                background: #e9ecef;
                padding: 2px 6px;
                border-radius: 4px;
                margin: 5px 0;
            }
            .status-200 { color: #28a745; }
            .status-404 { color: #dc3545; }
            .status-301, .status-302 { color: #ffc107; }
            .chart-container {
                margin: 30px 0;
                padding: 20px;
                background: #f8f9fa;
                border-radius: 10px;
            }
            .hour-bar {
                display: flex;
                align-items: center;
                margin: 5px 0;
            }
            .hour-label {
                min-width: 60px;
                font-weight: bold;
            }
            .hour-value {
                flex: 1;
                background: linear-gradient(90deg, #667eea, #764ba2);
                padding: 5px 10px;
                border-radius: 5px;
                color: white;
                font-size: 0.9em;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }
            th {
                background: linear-gradient(135deg, #667eea, #764ba2);
                color: white;
                padding: 12px;
                text-align: left;
            }
            td {
                padding: 10px;
                border-bottom: 1px solid #ddd;
            }
            tr:hover {
                background: #f5f7fa;
            }
            .top-pages {
                list-style: none;
                padding: 0;
            }
            .top-pages li {
                background: #f8f9fa;
                margin: 10px 0;
                padding: 10px;
                border-radius: 5px;
                border-left: 4px solid #667eea;
            }
            .timestamp {
                color: #888;
                font-size: 0.85em;
            }
            @media (max-width: 768px) {
                .timeline-item {
                    flex-direction: column;
                }
                .timeline-content {
                    margin-left: 0;
                    margin-top: 10px;
                }
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🤖 Bot Timeline Analysis Report</h1>
            <p style="text-align: center; color: #888; margin-bottom: 30px;">
                Generated: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """
            </p>
    """

        # خلاصه آماری
        html_content += """
            <h2>📊 Summary Statistics</h2>
            <div class="summary-grid">
        """

        # شمارش بات‌ها بر اساس دسته
        categories_stats = {
            'Search Engines': len(timeline['search_engines']),
            'AI Bots': len(timeline['ai_bots']),
            'Social Media': len(timeline['social_media']),
            'SEO Tools': len(timeline['seo_tools']),
            'Other Bots': len(timeline['other_bots'])
        }

        for category, count in categories_stats.items():
            total_requests = sum(len(events) for events in timeline[category.lower().replace(' ', '_')].values())
            html_content += f"""
                <div class="summary-card">
                    <h3>{category}</h3>
                    <p><strong>{count}</strong> unique bots</p>
                    <p><strong>{total_requests}</strong> total requests</p>
                </div>
            """

        html_content += "</div>"

        # Timeline اصلی
        html_content += """
            <h2>⏰ Complete Timeline</h2>
            <div class="timeline">
        """

        # نمایش آخرین 100 رویداد
        for event in timeline['timeline_events'][-100:]:
            bot_class = ""
            if 'google' in event['bot_name'].lower():
                bot_class = "bot-google"
            elif 'bing' in event['bot_name'].lower():
                bot_class = "bot-bing"
            elif 'openai' in event['bot_name'].lower():
                bot_class = "bot-openai"
            elif 'perplexity' in event['bot_name'].lower():
                bot_class = "bot-perplexity"
            elif 'meta' in event['bot_name'].lower() or 'facebook' in event['bot_name'].lower():
                bot_class = "bot-meta"

            status_class = f"status-{event['status_code']}"

            html_content += f"""
                <div class="timeline-item">
                    <div class="timeline-time">{event['datetime_obj'].strftime('%H:%M:%S')}</div>
                    <div class="timeline-content">
                        <span class="bot-badge {bot_class}">{event['bot_type']}</span>
                        <span class="timestamp">{event['datetime_obj'].strftime('%Y-%m-%d')}</span>
                        <div class="url-path">{event['url'][:100]}</div>
                        <div>
                            <small>IP: {event['ip']} | Status: <span class="{status_class}">{event['status_code']}</span></small>
                        </div>
                    </div>
                </div>
            """

        html_content += "</div>"

        # بخش تحلیل بات‌های AI
        html_content += """
            <h2>🤖 AI Bots Activity</h2>
            <table>
                <thead>
                    <tr>
                        <th>Bot Name</th>
                        <th>Total Requests</th>
                        <th>First Visit</th>
                        <th>Last Visit</th>
                        <th>Top Pages</th>
                    </tr>
                </thead>
                <tbody>
        """

        for bot_name, events in timeline['ai_bots'].items():
            if events:
                first_visit = min(e['datetime_obj'] for e in events)
                last_visit = max(e['datetime_obj'] for e in events)

                # Top pages
                page_counts = {}
                for e in events:
                    page_counts[e['url']] = page_counts.get(e['url'], 0) + 1
                top_pages = sorted(page_counts.items(), key=lambda x: x[1], reverse=True)[:3]

                html_content += f"""
                    <tr>
                        <td><strong>{bot_name}</strong></td>
                        <td>{len(events)}</td>
                        <td>{first_visit.strftime('%Y-%m-%d %H:%M')}</td>
                        <td>{last_visit.strftime('%Y-%m-%d %H:%M')}</td>
                        <td>{'<br>'.join([f"{p[0][:50]} ({p[1]}x)" for p in top_pages])}</td>
                    </tr>
                """

        html_content += """
                </tbody>
            </table>
        """

        # بخش تحلیل موتورهای جستجو
        html_content += """
            <h2>🔍 Search Engines Activity</h2>
            <table>
                <thead>
                    <tr>
                        <th>Search Engine</th>
                        <th>Bot Types</th>
                        <th>Total Requests</th>
                        <th>Date Range</th>
                        <th>Crawl Rate</th>
                    </tr>
                </thead>
                <tbody>
        """

        for bot_name, events in timeline['search_engines'].items():
            if events:
                # شناسایی انواع مختلف بات
                bot_types = list(set(e['bot_type'] for e in events))
                first_visit = min(e['datetime_obj'] for e in events)
                last_visit = max(e['datetime_obj'] for e in events)

                # محاسبه نرخ crawl
                time_diff = (last_visit - first_visit).total_seconds() / 60  # به دقیقه
                crawl_rate = len(events) / time_diff if time_diff > 0 else 0

                html_content += f"""
                    <tr>
                        <td><strong>{bot_name}</strong></td>
                        <td>{', '.join(bot_types[:3])}</td>
                        <td>{len(events)}</td>
                        <td>{first_visit.strftime('%Y-%m-%d')} to {last_visit.strftime('%Y-%m-%d')}</td>
                        <td>{crawl_rate:.1f} req/min</td>
                    </tr>
                """

        html_content += """
                </tbody>
            </table>
        """

        # نمودار فعالیت ساعتی
        html_content += """
            <h2>📈 Hourly Activity Pattern</h2>
            <div class="chart-container">
        """

        # جمع‌آوری آمار ساعتی برای امروز
        today = datetime.now().strftime('%Y-%m-%d')
        for hour in range(24):
            hour_key = f"{today} {hour:02d}:00"
            if hour_key in timeline['hourly_summary']:
                total = sum(timeline['hourly_summary'][hour_key].values())
                if total > 0:
                    html_content += f"""
                        <div class="hour-bar">
                            <div class="hour-label">{hour:02d}:00</div>
                            <div class="hour-value" style="width: {min(total*5, 100)}%">
                                {total} requests
                            </div>
                        </div>
                    """

        html_content += """
            </div>
        </div>
    </body>
    </html>
        """

        # ذخیره فایل HTML
        with open('bot_timeline_report.html', 'w', encoding='utf-8') as f:
            f.write(html_content)

        print("✅ گزارش HTML Timeline در bot_timeline_report.html ذخیره شد")    

    def _generate_timeline_text(self, timeline: Dict):
        """تولید گزارش Text از Timeline"""
        with open('bot_timeline_report.txt', 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("                    BOT TIMELINE ANALYSIS REPORT\n")
            f.write(f"                    Generated: {datetime.now()}\n")
            f.write("="*80 + "\n\n")

            # بخش موتورهای جستجو
            f.write("🔍 SEARCH ENGINES TIMELINE\n")
            f.write("-"*60 + "\n\n")

            for bot_name, events in timeline['search_engines'].items():
                if events:
                    f.write(f"### {bot_name}\n")
                    f.write(f"Total Visits: {len(events)}\n")

                    # گروه‌بندی بر اساس نوع بات
                    bot_types = {}
                    for e in events:
                        bot_type = e['bot_type']
                        if bot_type not in bot_types:
                            bot_types[bot_type] = []
                        bot_types[bot_type].append(e)

                    for bot_type, type_events in bot_types.items():
                        f.write(f"\n  [{bot_type}] - {len(type_events)} requests\n")

                        # نمایش آخرین 5 بازدید
                        for event in type_events[-5:]:
                            f.write(f"    • {event['datetime_obj'].strftime('%Y-%m-%d %H:%M:%S')}")
                            f.write(f" | {event['url'][:60]}")
                            f.write(f" | Status: {event['status_code']}")
                            f.write(f" | IP: {event['ip']}\n")

                    f.write("\n")

            # بخش بات‌های AI
            f.write("\n🤖 AI BOTS TIMELINE\n")
            f.write("-"*60 + "\n\n")

            for bot_name, events in timeline['ai_bots'].items():
                if events:
                    f.write(f"### {bot_name}\n")
                    f.write(f"Total Visits: {len(events)}\n")

                    first_visit = min(e['datetime_obj'] for e in events)
                    last_visit = max(e['datetime_obj'] for e in events)

                    f.write(f"First Visit: {first_visit.strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"Last Visit: {last_visit.strftime('%Y-%m-%d %H:%M:%S')}\n")

                    # تحلیل صفحات
                    page_counts = {}
                    for e in events:
                        page_counts[e['url']] = page_counts.get(e['url'], 0) + 1

                    top_pages = sorted(page_counts.items(), key=lambda x: x[1], reverse=True)[:5]
                    f.write("\nTop Crawled Pages:\n")
                    for page, count in top_pages:
                        f.write(f"  • {page[:80]} ({count}x)\n")

                    # نمایش timeline
                    f.write("\nRecent Activity:\n")
                    for event in events[-10:]:
                        f.write(f"  {event['datetime_obj'].strftime('%H:%M:%S')} - {event['url'][:60]}\n")

                    f.write("\n")

            # خلاصه روزانه
            f.write("\n📅 DAILY SUMMARY\n")
            f.write("-"*60 + "\n\n")

            for day, bots in sorted(timeline['daily_summary'].items())[-7:]:
                f.write(f"{day}:\n")
                for bot, count in sorted(bots.items(), key=lambda x: x[1], reverse=True):
                    f.write(f"  • {bot}: {count} requests\n")
                f.write("\n")

            # پرترافیک‌ترین صفحات
            f.write("\n📄 MOST CRAWLED PAGES BY BOTS\n")
            f.write("-"*60 + "\n\n")

            all_pages = {}
            for bot, pages in timeline['bot_page_analysis'].items():
                for page, visits in pages.items():
                    if page not in all_pages:
                        all_pages[page] = {'total': 0, 'bots': set()}
                    all_pages[page]['total'] += len(visits)
                    all_pages[page]['bots'].add(bot)

            top_pages_sorted = sorted(all_pages.items(), 
                                      key=lambda x: x[1]['total'], 
                                      reverse=True)[:20]

            for page, data in top_pages_sorted:
                f.write(f"{page[:80]}\n")
                f.write(f"  Total: {data['total']} | Bots: {', '.join(list(data['bots'])[:5])}\n\n")

        print("✅ گزارش Text Timeline در bot_timeline_report.txt ذخیره شد") 

    def _generate_timeline_json(self, timeline: Dict):
        """تولید گزارش JSON از Timeline"""
        # تبدیل datetime objects به string برای JSON
        json_timeline = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_events': len(timeline['timeline_events'])
            },
            'search_engines': {},
            'ai_bots': {},
            'social_media': {},
            'seo_tools': {},
            'other_bots': {},
            'hourly_summary': dict(timeline['hourly_summary']),
            'daily_summary': dict(timeline['daily_summary'])
        }

        # تبدیل events
        for category in ['search_engines', 'ai_bots', 'social_media', 'seo_tools', 'other_bots']:
            for bot_name, events in timeline[category].items():
                if events:
                    json_timeline[category][bot_name] = {
                        'total_requests': len(events),
                        'first_visit': min(e['datetime_obj'] for e in events).isoformat(),
                        'last_visit': max(e['datetime_obj'] for e in events).isoformat(),
                        'recent_activity': [
                            {
                                'timestamp': e['timestamp'],
                                'url': e['url'],
                                'status': e['status_code'],
                                'ip': e['ip']
                            }
                            for e in events[-20:]  # آخرین 20 رویداد
                        ]
                    }

        with open('bot_timeline.json', 'w', encoding='utf-8') as f:
            json.dump(json_timeline, f, indent=2, ensure_ascii=False)

        print("✅ گزارش JSON Timeline در bot_timeline.json ذخیره شد")   
    
    def export_all_reports(self):
        """تولید همه گزارش‌ها با رفع مشکل encoding"""
        print("\n📁 تولید همه گزارش‌ها...")
        
        # 1. Excel Report
        excel_file = self.export_to_excel(self.analysis_results)
        
        # 2. JSON Report
        json_file = self.export_json_report()
        
        # 3. Firewall Rules
        self.export_firewall_rules()

        # اضافه کردن گزارش Timeline
        timeline_report = self.generate_bot_timeline_report()

        # 4. Ban List - با encoding UTF-8
        with open('ban_list.txt', 'w', encoding='utf-8') as f:
            f.write(f"# Suspicious IPs - Generated: {datetime.now()}\n")
            f.write(f"# Total: {len(self.suspicious_ips)} IPs\n\n")
            
            # Sort by risk score
            sorted_ips = sorted(
                self.suspicious_ips,
                key=lambda ip: self.analysis_results['risk_scores'].get(ip, {}).get('score', 0),
                reverse=True
            )
            
            for ip in sorted_ips:
                risk_info = self.analysis_results['risk_scores'].get(ip, {})
                reasons_str = ', '.join(risk_info.get('reasons', ['Unknown'])[:2])
                f.write(f"{ip} # Score: {risk_info.get('score', 0)}, Level: {risk_info.get('risk_level', 'Unknown')}\n")
        
        # 5. Critical IPs for immediate action - با encoding UTF-8
        with open('critical_ips.txt', 'w', encoding='utf-8') as f:
            f.write(f"# CRITICAL IPs requiring immediate action\n")
            f.write(f"# Generated: {datetime.now()}\n\n")
            
            for ip in self.critical_ips:
                risk_info = self.analysis_results['risk_scores'].get(ip, {})
                reasons_str = ', '.join(risk_info.get('reasons', ['Unknown'])[:2])
                f.write(f"{ip} # {reasons_str}\n")
        
        # 6. Bot Visit Report - گزارش جدید زمان بازدید بات‌ها
        bot_visits = self.analyze_bot_visit_times()
        with open('bot_visits_report.txt', 'w', encoding='utf-8') as f:
            f.write(f"# Bot Visit Times Analysis\n")
            f.write(f"# Generated: {datetime.now()}\n\n")
            
            f.write("## LEGITIMATE BOTS\n")
            f.write("-" * 50 + "\n\n")
            
            for company, data in bot_visits['legitimate_bot_times'].items():
                if data['total_visits'] > 0:
                    f.write(f"### {company}\n")
                    f.write(f"Total Visits: {data['total_visits']}\n")
                    f.write(f"First Visit: {data['first_visit']}\n")
                    f.write(f"Last Visit: {data['last_visit']}\n")
                    f.write(f"Peak Hours: {', '.join([f'{h:02d}:00' for h in data['peak_hours']])}\n")
                    f.write(f"Unique IPs: {len(data['unique_ips'])}\n")
                    f.write(f"IP List: {', '.join(list(data['unique_ips'])[:10])}\n")
                    f.write("\n")
            
            if self.analysis_results['bot_analysis']['fake']:
                f.write("\n## FAKE BOTS DETECTED\n")
                f.write("-" * 50 + "\n\n")
                for ip, info in self.analysis_results['bot_analysis']['fake'].items():
                    f.write(f"IP: {ip}\n")
                    f.write(f"Claimed to be: {info.get('claimed', 'Unknown')}\n")
                    f.write(f"Reason: {info.get('reason', 'Failed verification')}\n\n")
        
        # 7. Summary Report (Markdown) - با encoding UTF-8
        with open('security_summary.md', 'w', encoding='utf-8') as f:
            f.write(f"# Security Analysis Report\n")
            f.write(f"**Generated:** {datetime.now()}\n")
            f.write(f"**Site Type:** {self.site_type.upper()}\n\n")
            
            f.write("## 📊 Overview\n")
            overview = self.analysis_results['overview']
            f.write(f"- Total Requests: {overview['total_requests']:,}\n")
            f.write(f"- Unique IPs: {overview['unique_ips']:,}\n")
            f.write(f"- Suspicious IPs: {len(self.suspicious_ips)}\n")
            f.write(f"- Critical Threats: {len(self.critical_ips)}\n")
            f.write(f"- Error Rate: {overview['error_rate']:.2f}%\n\n")
            
            f.write("## 🤖 Bot Analysis\n")
            bot_visits = self.analyze_bot_visit_times()
            for company, data in bot_visits['legitimate_bot_times'].items():
                if data['total_visits'] > 0:
                    f.write(f"### {company}\n")
                    f.write(f"- Visits: {data['total_visits']}\n")
                    f.write(f"- Time Range: {data['first_visit']} to {data['last_visit']}\n")
                    f.write(f"- Peak Hours: {', '.join([f'{h:02d}:00' for h in data['peak_hours'][:3]])}\n\n")
            
            f.write("## 🚨 Critical Threats\n")
            for ip in list(self.critical_ips)[:10]:
                risk_info = self.analysis_results['risk_scores'][ip]
                f.write(f"### {ip}\n")
                f.write(f"- Risk Score: {risk_info['score']}\n")
                f.write(f"- Reasons: {', '.join(risk_info['reasons'][:3])}\n\n")
            
            f.write("## 🎯 Attack Types Detected\n")
            for attack_type, ip_dict in self.analysis_results['attack_analysis'].items():
                if ip_dict:
                    f.write(f"- **{attack_type.replace('_', ' ').title()}**: {len(ip_dict)} IPs\n")
            
        
        print("\n✅ همه گزارش‌ها با موفقیت تولید شدند:")
        print(f"  📊 Excel: {excel_file}")
        print(f"  📄 JSON: {json_file}")
        print("  🔒 Firewall Rules: iptables_rules.sh, htaccess_rules.txt, nginx_rules.conf")
        print("  📝 Ban Lists: ban_list.txt, critical_ips.txt")
        print("  🤖 Bot Report: bot_visits_report.txt")
        print("  📑 Summary: security_summary.md")


def main():
    """تابع اصلی برنامه"""
    import sys
    import argparse
    from datetime import timedelta
    
    # Parser برای آرگومان‌های خط فرمان
    parser = argparse.ArgumentParser(
        description='Advanced Security Log Analyzer',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python analyzer.py access_log.gz --type wordpress --period 30
  python analyzer.py logs.zip --type opencart --excel --period 90
  python analyzer.py access.log --type general --all
        """
    )
    
    parser.add_argument('logfile', help='Path to log file (.gz, .zip, or plain text)')
    parser.add_argument('--type', '-t', 
                       choices=['wordpress', 'opencart', 'general'],
                       default='general',
                       help='Type of website (default: general)')
    parser.add_argument('--period', '-p',
                       type=int,
                       choices=[30, 60, 90, 180, 365, 0],
                       help='Time period in days (30/60/90/180/365/0 for all)')
    parser.add_argument('--interactive', '-i',
                       action='store_true',
                       help='Interactive mode for time period selection')
    parser.add_argument('--excel', '-e', 
                       action='store_true',
                       help='Generate Excel report')
    parser.add_argument('--json', '-j',
                       action='store_true',
                       help='Generate JSON report')
    parser.add_argument('--firewall', '-f',
                       action='store_true',
                       help='Generate firewall rules')
    parser.add_argument('--all', '-a',
                       action='store_true',
                       help='Generate all reports')
    parser.add_argument('--quiet', '-q',
                       action='store_true',
                       help='Minimal output')
    parser.add_argument('--timeline', '-tl',
                       action='store_true',
                       help='Generate bot timeline report')

    args = parser.parse_args()
    
    # بررسی وجود فایل
    if not os.path.exists(args.logfile):
        print(f"❌ فایل {args.logfile} یافت نشد")
        sys.exit(1)
    
    # Header
    if not args.quiet:
        print("\n" + "="*80)
        print("🔒 Advanced Security Log Analyzer v2.0")
        print("="*80)
        print(f"📁 فایل: {args.logfile}")
        print(f"🌐 نوع سایت: {args.type}")
    
    try:
        # ایجاد آنالایزر
        analyzer = AdvancedSecurityAnalyzer(args.logfile, site_type=args.type)
        
        # انتخاب بازه زمانی
        days_limit = 0
        
        if args.interactive or (not args.period and not args.quiet):
            # حالت interactive: نمایش منو
            days_limit = analyzer.select_time_period()
        elif args.period is not None:
            # استفاده از مقدار خط فرمان
            days_limit = args.period
            period_names = {
                30: 'یک ماه اخیر',
                60: 'دو ماه اخیر',
                90: 'سه ماه اخیر',
                180: 'شش ماه اخیر',
                365: 'دوازده ماه اخیر',
                0: 'کل لاگ‌ها'
            }
            if not args.quiet:
                print(f"\n📅 بازه زمانی: {period_names.get(days_limit, f'{days_limit} روز')}")
        
        # بارگذاری لاگ‌ها با فیلتر زمانی
        if not analyzer.load_logs(days_limit):
            print("❌ خطا در بارگذاری فایل لاگ")
            sys.exit(1)
        
        # بررسی که آیا لاگی در بازه زمانی وجود دارد
        if not analyzer.logs:
            print("\n⚠️ هیچ لاگی در بازه زمانی انتخابی یافت نشد")
            print("💡 لطفاً بازه زمانی بزرگتری انتخاب کنید یا از گزینه 'کل لاگ‌ها' استفاده کنید")
            sys.exit(1)
        
        # ادامه تحلیل...
        analyzer.generate_report()
        
        # تولید خروجی‌های درخواستی
        if args.all:
            analyzer.export_all_reports()
        else:
            if args.excel:
                analyzer.export_to_excel(analyzer.analysis_results)
            if args.json:
                analyzer.export_json_report()
            if args.firewall:
                analyzer.export_firewall_rules()
            if args.timeline:
                analyzer.generate_bot_timeline_report()

            # همیشه لیست بن را تولید کن
            with open('ban_list.txt', 'w', encoding='utf-8') as f:
                f.write(f"# Suspicious IPs - {datetime.now()}\n")
                if days_limit > 0:
                    f.write(f"# Time Period: Last {days_limit} days\n")
                for ip in sorted(analyzer.suspicious_ips):
                    f.write(f"{ip}\n")
            print(f"\n✅ لیست IP های مشکوک در ban_list.txt ذخیره شد")
        
        # خلاصه نهایی
        if not args.quiet:
            print("\n" + "="*80)
            print("📊 خلاصه نهایی:")
            print(f"  • {len(analyzer.suspicious_ips)} IP مشکوک شناسایی شد")
            print(f"  • {len(analyzer.critical_ips)} IP بحرانی نیاز به اقدام فوری")
            print(f"  • {len(analyzer.analysis_results['bot_analysis']['fake'])} بات جعلی")
            
            if analyzer.critical_ips:
                print("\n⚠️ هشدار: IP های زیر باید فوراً بن شوند:")
                for ip in list(analyzer.critical_ips)[:5]:
                    print(f"  🔴 {ip}")
            
            print("\n💡 برای مشاهده جزئیات، فایل‌های گزارش را بررسی کنید")
    
    except KeyboardInterrupt:
        print("\n\n⚠️ عملیات توسط کاربر لغو شد")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ خطای غیرمنتظره: {e}")
        import traceback
        if not args.quiet:
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
